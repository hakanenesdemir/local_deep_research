import time
import psutil
import os
import json
import glob
from typing import Dict, List, Optional, Tuple
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import numpy as np
from sentence_transformers import SentenceTransformer
from transformers import AutoTokenizer, AutoModel, AutoModelForQuestionAnswering
import torch
import warnings
warnings.filterwarnings('ignore')

# Database imports
import lancedb
import chromadb
from qdrant_client import QdrantClient
from qdrant_client.models import Distance, VectorParams, PointStruct, SearchParams
from pymilvus import MilvusClient
import weaviate
from weaviate.classes.config import Property, DataType, Configure


class VectorDatabaseBenchmark:
    """
    Models klasöründeki verileri okuyup tüm veritabanlarına yazan
    ve benchmark testleri yapan kapsamlı sınıf.
    """
    
    def __init__(self, models_path: str = None, db_base_path: str = None):
        self.base_path = "/home/ugo/Documents/Python/bitirememe projesi"
        self.models_path = models_path or os.path.join(self.base_path, "models")
        self.db_base_path = db_base_path or os.path.join(self.base_path, "DB")
        
        # Model bilgileri
        self.model_info = {
            "embedding_model": {
                "name": "all-MiniLM-L6-v2",
                "type": "SentenceTransformer",
                "vector_dim": 384,
                "description": "Hafif embedding modeli"
            },
            "deberta_qa_model": {
                "name": "deberta_v3_qa_model",
                "type": "DeBERTa-v3-QA",
                "path": os.path.join(self.models_path, "deberta_v3_qa_model"),
                "description": "DeBERTa v3 Soru-Cevap modeli"
            },
            "electra_qa_model": {
                "name": "electra_english_qa_model",
                "type": "ELECTRA-QA",
                "path": os.path.join(self.models_path, "electra_english_qa_model"),
                "description": "ELECTRA Soru-Cevap modeli"
            },
            "qa_model": {
                "name": "qa_model",
                "type": "QA-Model",
                "path": os.path.join(self.models_path, "qa_model"),
                "description": "Genel Soru-Cevap modeli"
            },
            "xlm_roberta_qa_model": {
                "name": "xlm_roberta_qa_model",
                "type": "XLM-RoBERTa-QA",
                "path": os.path.join(self.models_path, "xlm_roberta_qa_model"),
                "description": "XLM-RoBERTa çok dilli model"
            }
        }
        
        # Modelleri yükle
        print("="*70)
        print("📊 MODELLER YÜKLENİYOR")
        print("="*70)
        
        # Model storage
        self.models = {}
        self.tokenizers = {}
        self.vector_dims = {}
        self.all_embeddings = {}  # Her model için ayrı embedding
        self.all_query_vectors = {}  # Her model için ayrı sorgu vektörleri
        
        # 1. SentenceTransformer
        print(f"\n🔹 Embedding Modeli: all-MiniLM-L6-v2")
        self.embedding_model = SentenceTransformer("all-MiniLM-L6-v2")
        self.vector_dim = 384
        self.models['sentence_transformer'] = self.embedding_model
        self.vector_dims['sentence_transformer'] = 384
        self.model_info['embedding_model']['status'] = 'loaded'
        print("   ✓ Yüklendi (dim: 384)")
        
        # 2-5. Diğer modeller
        for model_key in ['deberta_qa_model', 'electra_qa_model', 'qa_model', 'xlm_roberta_qa_model']:
            self._load_model(model_key)
        
        # Test sorguları
        self.test_queries = [
            "artificial intelligence healthcare applications",
            "machine learning medical diagnosis systems",
            "deep learning neural network architectures",
            "natural language processing techniques",
            "computer vision medical imaging analysis",
            "reinforcement learning robotics control",
            "transformer models attention mechanism",
            "convolutional neural networks image classification",
            "recurrent neural networks sequence modeling",
            "generative adversarial networks image synthesis"
        ]
        
        # Sonuçlar
        self.results = {
            "write_benchmark": {},
            "search_benchmark": {},
            "model_benchmark": {},
            "multi_model_benchmark": {},
            "metadata": {
                "date": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                "query_count": len(self.test_queries),
                "models": self.model_info
            }
        }
        
        self.documents = []
    
    def _load_model(self, model_key: str):
        """Model yükle ve sakla"""
        fallback_models = {
            'deberta_qa_model': 'microsoft/deberta-v3-base',
            'electra_qa_model': 'google/electra-base-discriminator',
            'qa_model': 'bert-base-uncased',
            'xlm_roberta_qa_model': 'xlm-roberta-base'
        }
        
        model_path = self.model_info[model_key]['path']
        model_name = self.model_info[model_key]['name']
        fallback = fallback_models.get(model_key, 'bert-base-uncased')
        
        print(f"\n🔹 Model: {model_name}")
        print(f"   Yol: {model_path}")
        
        try:
            if os.path.exists(model_path) and os.path.exists(os.path.join(model_path, "model.safetensors")):
                print(f"   📂 Model dosyası bulundu")
                
                try:
                    tokenizer = AutoTokenizer.from_pretrained(model_path)
                except:
                    tokenizer = AutoTokenizer.from_pretrained(fallback)
                
                try:
                    model = AutoModelForQuestionAnswering.from_pretrained(model_path)
                    self.model_info[model_key]['model_type'] = 'QuestionAnswering'
                except:
                    try:
                        model = AutoModel.from_pretrained(model_path)
                        self.model_info[model_key]['model_type'] = 'BaseModel'
                    except:
                        model = AutoModel.from_pretrained(fallback)
                        self.model_info[model_key]['model_type'] = 'Fallback'
                
                vector_dim = model.config.hidden_size
                self.models[model_key] = model
                self.tokenizers[model_key] = tokenizer
                self.vector_dims[model_key] = vector_dim
                self.model_info[model_key]['vector_dim'] = vector_dim
                self.model_info[model_key]['status'] = 'loaded'
                print(f"   ✓ Yüklendi (dim: {vector_dim})")
            else:
                print(f"   ⚠ Model bulunamadı")
                self.model_info[model_key]['status'] = 'not_found'
        except Exception as e:
            print(f"   ❌ Hata: {e}")
            self.model_info[model_key]['status'] = 'error'
    
    def _get_model_embeddings(self, model_key: str, texts: List[str]) -> List[List[float]]:
        """Belirli bir model ile embedding oluştur"""
        if model_key == 'sentence_transformer':
            return self.embedding_model.encode(texts, show_progress_bar=False).tolist()
        
        model = self.models.get(model_key)
        tokenizer = self.tokenizers.get(model_key)
        
        if model is None or tokenizer is None:
            return None
        
        embeddings = []
        model.eval()
        
        with torch.no_grad():
            for text in texts:
                inputs = tokenizer(text, return_tensors="pt", truncation=True, max_length=512, padding=True)
                
                try:
                    if hasattr(model, 'deberta'):
                        outputs = model.deberta(**inputs)
                    elif hasattr(model, 'electra'):
                        outputs = model.electra(**inputs)
                    elif hasattr(model, 'roberta'):
                        outputs = model.roberta(**inputs)
                    elif hasattr(model, 'bert'):
                        outputs = model.bert(**inputs)
                    else:
                        outputs = model(**inputs, output_hidden_states=True)
                        if hasattr(outputs, 'hidden_states') and outputs.hidden_states:
                            outputs.last_hidden_state = outputs.hidden_states[-1]
                except:
                    outputs = model(**inputs)
                
                if hasattr(outputs, 'last_hidden_state'):
                    last_hidden = outputs.last_hidden_state
                else:
                    last_hidden = outputs[0]
                
                attention_mask = inputs['attention_mask']
                mask_expanded = attention_mask.unsqueeze(-1).expand(last_hidden.size()).float()
                sum_embeddings = torch.sum(last_hidden * mask_expanded, 1)
                sum_mask = torch.clamp(mask_expanded.sum(1), min=1e-9)
                embedding = (sum_embeddings / sum_mask).squeeze().tolist()
                embeddings.append(embedding)
        
        return embeddings
    
    def prepare_all_embeddings(self):
        """Tüm modeller için embedding'leri hazırla"""
        print("\n" + "="*70)
        print("📊 TÜM MODELLER İÇİN EMBEDDİNG'LER HAZIRLANIYOR")
        print("="*70)
        
        texts = [doc["text"] for doc in self.documents]
        
        for model_key in ['sentence_transformer', 'deberta_qa_model', 'electra_qa_model', 'qa_model', 'xlm_roberta_qa_model']:
            if model_key == 'sentence_transformer' or self.models.get(model_key) is not None:
                print(f"\n📊 {model_key} embedding'leri hesaplanıyor...")
                start_time = time.time()
                
                if model_key == 'sentence_transformer':
                    self.all_embeddings[model_key] = self.embedding_model.encode(texts, show_progress_bar=True).tolist()
                    self.all_query_vectors[model_key] = [self.embedding_model.encode(q).tolist() for q in self.test_queries]
                else:
                    # Batch processing for transformer models
                    embeddings = []
                    batch_size = 32
                    for i in range(0, len(texts), batch_size):
                        batch = texts[i:i+batch_size]
                        batch_emb = self._get_model_embeddings(model_key, batch)
                        if batch_emb:
                            embeddings.extend(batch_emb)
                        print(f"   {min(i+batch_size, len(texts))}/{len(texts)} işlendi...")
                    
                    if embeddings:
                        self.all_embeddings[model_key] = embeddings
                        self.all_query_vectors[model_key] = self._get_model_embeddings(model_key, self.test_queries)
                
                elapsed = time.time() - start_time
                print(f"   ✓ {model_key}: {elapsed:.2f}s ({len(texts)} doküman)")
            else:
                print(f"\n⚠ {model_key} yüklenmediği için atlanıyor")
    
    # ==================== MULTI-MODEL DATABASE WRITE ====================
    def write_all_models_to_milvus(self):
        """Tüm modeller için Milvus'a yaz"""
        print("\n" + "="*70)
        print("🔷 TÜM MODELLER İÇİN MILVUS'A YAZILIYOR")
        print("="*70)
        
        results = {}
        
        for model_key, embeddings in self.all_embeddings.items():
            if not embeddings:
                continue
                
            vector_dim = len(embeddings[0])
            collection_name = f"docs_{model_key}"
            db_path = os.path.join(self.db_base_path, f"milvus/{model_key}_db.db")
            os.makedirs(os.path.dirname(db_path), exist_ok=True)
            
            print(f"\n📝 {model_key} (dim: {vector_dim})...")
            
            try:
                if os.path.exists(db_path):
                    os.remove(db_path)
                
                client = MilvusClient(db_path)
                start_time = time.time()
                
                client.create_collection(collection_name=collection_name, dimension=vector_dim, metric_type="COSINE")
                
                data = [{"id": i, "vector": emb, "text": doc["text"][:500], "source": doc["source"]} 
                        for i, (doc, emb) in enumerate(zip(self.documents, embeddings))]
                
                for i in range(0, len(data), 100):
                    client.insert(collection_name=collection_name, data=data[i:i+100])
                
                write_time = time.time() - start_time
                results[model_key] = {
                    "status": "success",
                    "write_time": write_time,
                    "record_count": len(self.documents),
                    "vector_dim": vector_dim
                }
                print(f"   ✓ {write_time:.2f}s ({len(self.documents)} kayıt)")
                
            except Exception as e:
                results[model_key] = {"status": "error", "error": str(e)}
                print(f"   ❌ Hata: {e}")
        
        self.results["multi_model_benchmark"]["milvus_write"] = results
        return results
    
    def write_all_models_to_qdrant(self):
        """Tüm modeller için Qdrant'a yaz"""
        print("\n" + "="*70)
        print("🔷 TÜM MODELLER İÇİN QDRANT'A YAZILIYOR")
        print("="*70)
        
        results = {}
        
        try:
            client = QdrantClient(host="localhost", port=6333, timeout=120)
        except Exception as e:
            print(f"❌ Qdrant bağlantı hatası: {e}")
            return {"error": str(e)}
        
        for model_key, embeddings in self.all_embeddings.items():
            if not embeddings:
                continue
                
            vector_dim = len(embeddings[0])
            collection_name = f"docs_{model_key}"
            
            print(f"\n📝 {model_key} (dim: {vector_dim})...")
            
            try:
                try:
                    client.delete_collection(collection_name)
                except:
                    pass
                
                start_time = time.time()
                client.create_collection(collection_name=collection_name, vectors_config=VectorParams(size=vector_dim, distance=Distance.COSINE))
                
                points = [PointStruct(id=i, vector=emb, payload={"text": doc["text"], "source": doc["source"]}) 
                         for i, (doc, emb) in enumerate(zip(self.documents, embeddings))]
                
                for i in range(0, len(points), 100):
                    client.upsert(collection_name=collection_name, points=points[i:i+100])
                
                write_time = time.time() - start_time
                results[model_key] = {"status": "success", "write_time": write_time, "record_count": len(self.documents), "vector_dim": vector_dim}
                print(f"   ✓ {write_time:.2f}s")
                
            except Exception as e:
                results[model_key] = {"status": "error", "error": str(e)}
                print(f"   ❌ Hata: {e}")
        
        self.results["multi_model_benchmark"]["qdrant_write"] = results
        return results
    
    def write_all_models_to_chromadb(self):
        """Tüm modeller için ChromaDB'ye yaz"""
        print("\n" + "="*70)
        print("🔷 TÜM MODELLER İÇİN CHROMADB'YE YAZILIYOR")
        print("="*70)
        
        results = {}
        import shutil
        
        for model_key, embeddings in self.all_embeddings.items():
            if not embeddings:
                continue
                
            vector_dim = len(embeddings[0])
            collection_name = f"docs_{model_key}"
            db_path = os.path.join(self.db_base_path, f"chromadb/{model_key}_db")
            
            print(f"\n📝 {model_key} (dim: {vector_dim})...")
            
            try:
                if os.path.exists(db_path):
                    shutil.rmtree(db_path)
                os.makedirs(db_path, exist_ok=True)
                
                client = chromadb.PersistentClient(path=db_path)
                start_time = time.time()
                
                collection = client.create_collection(name=collection_name, metadata={"hnsw:space": "cosine"})
                
                ids = [str(i) for i in range(len(self.documents))]
                texts = [doc["text"] for doc in self.documents]
                metadatas = [{"source": doc["source"]} for doc in self.documents]
                
                for i in range(0, len(self.documents), 100):
                    end_idx = min(i + 100, len(self.documents))
                    collection.add(ids=ids[i:end_idx], embeddings=embeddings[i:end_idx], documents=texts[i:end_idx], metadatas=metadatas[i:end_idx])
                
                write_time = time.time() - start_time
                results[model_key] = {"status": "success", "write_time": write_time, "record_count": len(self.documents), "vector_dim": vector_dim}
                print(f"   ✓ {write_time:.2f}s")
                
            except Exception as e:
                results[model_key] = {"status": "error", "error": str(e)}
                print(f"   ❌ Hata: {e}")
        
        self.results["multi_model_benchmark"]["chromadb_write"] = results
        return results
    
    def write_all_models_to_lancedb(self):
        """Tüm modeller için LanceDB'ye yaz"""
        print("\n" + "="*70)
        print("🔷 TÜM MODELLER İÇİN LANCEDB'YE YAZILIYOR")
        print("="*70)
        
        results = {}
        import shutil
        
        for model_key, embeddings in self.all_embeddings.items():
            if not embeddings:
                continue
                
            vector_dim = len(embeddings[0])
            table_name = f"docs_{model_key}"
            db_path = os.path.join(self.db_base_path, f"lancedb/{model_key}_db")
            
            print(f"\n📝 {model_key} (dim: {vector_dim})...")
            
            try:
                if os.path.exists(db_path):
                    shutil.rmtree(db_path)
                os.makedirs(db_path, exist_ok=True)
                
                db = lancedb.connect(db_path)
                start_time = time.time()
                
                data = [{"id": i, "vector": emb, "text": doc["text"], "source": doc["source"]} 
                       for i, (doc, emb) in enumerate(zip(self.documents, embeddings))]
                
                table = db.create_table(table_name, data=data)
                
                write_time = time.time() - start_time
                results[model_key] = {"status": "success", "write_time": write_time, "record_count": len(self.documents), "vector_dim": vector_dim}
                print(f"   ✓ {write_time:.2f}s")
                
            except Exception as e:
                results[model_key] = {"status": "error", "error": str(e)}
                print(f"   ❌ Hata: {e}")
        
        self.results["multi_model_benchmark"]["lancedb_write"] = results
        return results
    
    def write_all_models_to_weaviate(self):
        """Tüm modeller için Weaviate'a yaz"""
        print("\n" + "="*70)
        print("🔷 TÜM MODELLER İÇİN WEAVIATE'A YAZILIYOR")
        print("="*70)
        
        results = {}
        
        try:
            client = weaviate.connect_to_local()
        except Exception as e:
            print(f"❌ Weaviate bağlantı hatası: {e}")
            return {"error": str(e)}
        
        for model_key, embeddings in self.all_embeddings.items():
            if not embeddings:
                continue
                
            vector_dim = len(embeddings[0])
            # Weaviate collection name format
            collection_name = f"Docs{model_key.replace('_', '').title()}"
            
            print(f"\n📝 {model_key} (dim: {vector_dim})...")
            
            try:
                try:
                    client.collections.delete(collection_name)
                except:
                    pass
                
                start_time = time.time()
                
                collection = client.collections.create(
                    name=collection_name,
                    vectorizer_config=Configure.Vectorizer.none(),
                    properties=[
                        Property(name="text", data_type=DataType.TEXT),
                        Property(name="source", data_type=DataType.TEXT)
                    ]
                )
                
                with collection.batch.dynamic() as batch:
                    for i, (doc, emb) in enumerate(zip(self.documents, embeddings)):
                        batch.add_object(properties={"text": doc["text"], "source": doc["source"]}, vector=emb)
                
                write_time = time.time() - start_time
                results[model_key] = {"status": "success", "write_time": write_time, "record_count": len(self.documents), "vector_dim": vector_dim}
                print(f"   ✓ {write_time:.2f}s")
                
            except Exception as e:
                results[model_key] = {"status": "error", "error": str(e)}
                print(f"   ❌ Hata: {e}")
        
        self.results["multi_model_benchmark"]["weaviate_write"] = results
        return results
    
    # ==================== MULTI-MODEL SEARCH BENCHMARK ====================
    def benchmark_all_models_milvus_search(self):
        """Tüm modeller için Milvus arama benchmark - Çoklu algoritma"""
        print("\n" + "="*70)
        print("🔍 TÜM MODELLER İÇİN MILVUS ARAMA BENCHMARK")
        print("="*70)
        
        results = {}
        
        for model_key, query_vectors in self.all_query_vectors.items():
            if not query_vectors:
                continue
            
            collection_name = f"docs_{model_key}"
            db_path = os.path.join(self.db_base_path, f"milvus/{model_key}_db.db")
            
            print(f"\n📊 {model_key}...")
            
            try:
                client = MilvusClient(db_path)
                model_results = {}
                
                # 1. Default HNSW Search
                def hnsw_search():
                    return [client.search(collection_name=collection_name, data=[qv], limit=10, output_fields=["text"]) for qv in query_vectors]
                
                perf = self._measure_search_time(hnsw_search)
                if "error" not in perf:
                    model_results["HNSW_default"] = {"performance": perf, "qps": len(self.test_queries) / perf["avg_time"]}
                    print(f"   ✓ HNSW_default: {perf['avg_time']*1000:.2f}ms")
                
                # 2. Farklı limit değerleri
                for limit in [5, 20, 50, 100]:
                    def limit_search(l=limit):
                        return [client.search(collection_name=collection_name, data=[qv], limit=l, output_fields=["text"]) for qv in query_vectors]
                    
                    perf = self._measure_search_time(limit_search)
                    if "error" not in perf:
                        model_results[f"HNSW_limit{limit}"] = {"performance": perf, "qps": len(self.test_queries) / perf["avg_time"]}
                        print(f"   ✓ HNSW_limit{limit}: {perf['avg_time']*1000:.2f}ms")
                
                # 3. Batch search (tüm sorguları tek seferde)
                def batch_search():
                    return client.search(collection_name=collection_name, data=query_vectors, limit=10, output_fields=["text"])
                
                perf = self._measure_search_time(batch_search)
                if "error" not in perf:
                    model_results["HNSW_batch"] = {"performance": perf, "qps": len(self.test_queries) / perf["avg_time"]}
                    print(f"   ✓ HNSW_batch: {perf['avg_time']*1000:.2f}ms")
                
                # 4. Farklı nprobe değerleri (search params)
                for nprobe in [1, 8, 16, 32]:
                    def nprobe_search(np=nprobe):
                        return [client.search(
                            collection_name=collection_name, 
                            data=[qv], 
                            limit=10, 
                            output_fields=["text"],
                            search_params={"nprobe": np}
                        ) for qv in query_vectors]
                    
                    perf = self._measure_search_time(nprobe_search)
                    if "error" not in perf:
                        model_results[f"HNSW_nprobe{nprobe}"] = {"performance": perf, "qps": len(self.test_queries) / perf["avg_time"]}
                        print(f"   ✓ HNSW_nprobe{nprobe}: {perf['avg_time']*1000:.2f}ms")
                
                results[model_key] = model_results
                    
            except Exception as e:
                results[model_key] = {"error": str(e)}
                print(f"   ❌ Hata: {e}")
        
        self.results["multi_model_benchmark"]["milvus_search"] = results
        return results
    
    def benchmark_all_models_qdrant_search(self):
        """Tüm modeller için Qdrant arama benchmark - Çoklu algoritma"""
        print("\n" + "="*70)
        print("🔍 TÜM MODELLER İÇİN QDRANT ARAMA BENCHMARK")
        print("="*70)
        
        results = {}
        
        try:
            client = QdrantClient(host="localhost", port=6333, timeout=60)
        except Exception as e:
            print(f"❌ Qdrant bağlantı hatası: {e}")
            return {"error": str(e)}
        
        for model_key, query_vectors in self.all_query_vectors.items():
            if not query_vectors:
                continue
            
            collection_name = f"docs_{model_key}"
            print(f"\n📊 {model_key}...")
            
            try:
                model_results = {}
                
                # 1. Default HNSW Search
                def hnsw_search():
                    return [client.query_points(collection_name=collection_name, query=qv, limit=10, with_payload=True) for qv in query_vectors]
                
                perf = self._measure_search_time(hnsw_search)
                if "error" not in perf:
                    model_results["HNSW_default"] = {"performance": perf, "qps": len(self.test_queries) / perf["avg_time"]}
                    print(f"   ✓ HNSW_default: {perf['avg_time']*1000:.2f}ms")
                
                # 2. Exact Search (brute force)
                def exact_search():
                    return [client.query_points(collection_name=collection_name, query=qv, limit=10, with_payload=True, search_params=SearchParams(exact=True)) for qv in query_vectors]
                
                perf = self._measure_search_time(exact_search)
                if "error" not in perf:
                    model_results["EXACT_bruteforce"] = {"performance": perf, "qps": len(self.test_queries) / perf["avg_time"]}
                    print(f"   ✓ EXACT_bruteforce: {perf['avg_time']*1000:.2f}ms")
                
                # 3. HNSW with different ef values
                for ef in [8, 16, 32, 64, 128, 256]:
                    def ef_search(ef_val=ef):
                        return [client.query_points(collection_name=collection_name, query=qv, limit=10, with_payload=True, search_params=SearchParams(hnsw_ef=ef_val)) for qv in query_vectors]
                    
                    perf = self._measure_search_time(ef_search)
                    if "error" not in perf:
                        model_results[f"HNSW_ef{ef}"] = {"performance": perf, "qps": len(self.test_queries) / perf["avg_time"]}
                        print(f"   ✓ HNSW_ef{ef}: {perf['avg_time']*1000:.2f}ms")
                
                # 4. Farklı limit değerleri
                for limit in [5, 20, 50, 100]:
                    def limit_search(l=limit):
                        return [client.query_points(collection_name=collection_name, query=qv, limit=l, with_payload=True) for qv in query_vectors]
                    
                    perf = self._measure_search_time(limit_search)
                    if "error" not in perf:
                        model_results[f"HNSW_limit{limit}"] = {"performance": perf, "qps": len(self.test_queries) / perf["avg_time"]}
                        print(f"   ✓ HNSW_limit{limit}: {perf['avg_time']*1000:.2f}ms")
                
                # 5. Quantization enabled search (if available)
                try:
                    def quantized_search():
                        return [client.query_points(
                            collection_name=collection_name, 
                            query=qv, 
                            limit=10, 
                            with_payload=True,
                            search_params=SearchParams(quantization={"ignore": False, "rescore": True})
                        ) for qv in query_vectors]
                    
                    perf = self._measure_search_time(quantized_search)
                    if "error" not in perf:
                        model_results["HNSW_quantized"] = {"performance": perf, "qps": len(self.test_queries) / perf["avg_time"]}
                        print(f"   ✓ HNSW_quantized: {perf['avg_time']*1000:.2f}ms")
                except:
                    pass
                
                # 6. Payload ile ve payload'sız arama karşılaştırması
                def no_payload_search():
                    return [client.query_points(collection_name=collection_name, query=qv, limit=10, with_payload=False) for qv in query_vectors]
                
                perf = self._measure_search_time(no_payload_search)
                if "error" not in perf:
                    model_results["HNSW_no_payload"] = {"performance": perf, "qps": len(self.test_queries) / perf["avg_time"]}
                    print(f"   ✓ HNSW_no_payload: {perf['avg_time']*1000:.2f}ms")
                
                results[model_key] = model_results
                
            except Exception as e:
                results[model_key] = {"error": str(e)}
                print(f"   ❌ Hata: {e}")
        
        self.results["multi_model_benchmark"]["qdrant_search"] = results
        return results
    
    def benchmark_all_models_chromadb_search(self):
        """Tüm modeller için ChromaDB arama benchmark - Çoklu algoritma"""
        print("\n" + "="*70)
        print("🔍 TÜM MODELLER İÇİN CHROMADB ARAMA BENCHMARK")
        print("="*70)
        
        results = {}
        
        for model_key, query_vectors in self.all_query_vectors.items():
            if not query_vectors:
                continue
            
            collection_name = f"docs_{model_key}"
            db_path = os.path.join(self.db_base_path, f"chromadb/{model_key}_db")
            
            print(f"\n📊 {model_key}...")
            
            try:
                client = chromadb.PersistentClient(path=db_path)
                collection = client.get_collection(collection_name)
                
                model_results = {}
                
                # 1. Default HNSW Vector Search
                def vector_search():
                    return [collection.query(query_embeddings=[qv], n_results=10) for qv in query_vectors]
                
                perf = self._measure_search_time(vector_search)
                if "error" not in perf:
                    model_results["HNSW_vector"] = {"performance": perf, "qps": len(self.test_queries) / perf["avg_time"]}
                    print(f"   ✓ HNSW_vector: {perf['avg_time']*1000:.2f}ms")
                
                # 2. Farklı n_results değerleri
                for n in [5, 20, 50, 100]:
                    def n_search(n_val=n):
                        return [collection.query(query_embeddings=[qv], n_results=n_val) for qv in query_vectors]
                    
                    perf = self._measure_search_time(n_search)
                    if "error" not in perf:
                        model_results[f"HNSW_n{n}"] = {"performance": perf, "qps": len(self.test_queries) / perf["avg_time"]}
                        print(f"   ✓ HNSW_n{n}: {perf['avg_time']*1000:.2f}ms")
                
                # 3. Text query search (eğer embedding_function varsa)
                def text_search():
                    return [collection.query(query_texts=[q], n_results=10) for q in self.test_queries]
                
                try:
                    perf = self._measure_search_time(text_search)
                    if "error" not in perf:
                        model_results["TEXT_search"] = {"performance": perf, "qps": len(self.test_queries) / perf["avg_time"]}
                        print(f"   ✓ TEXT_search: {perf['avg_time']*1000:.2f}ms")
                except:
                    pass
                
                # 4. Include/Exclude different fields
                def minimal_search():
                    return [collection.query(query_embeddings=[qv], n_results=10, include=["distances"]) for qv in query_vectors]
                
                perf = self._measure_search_time(minimal_search)
                if "error" not in perf:
                    model_results["HNSW_minimal"] = {"performance": perf, "qps": len(self.test_queries) / perf["avg_time"]}
                    print(f"   ✓ HNSW_minimal: {perf['avg_time']*1000:.2f}ms")
                
                def full_search():
                    return [collection.query(query_embeddings=[qv], n_results=10, include=["documents", "metadatas", "distances", "embeddings"]) for qv in query_vectors]
                
                perf = self._measure_search_time(full_search)
                if "error" not in perf:
                    model_results["HNSW_full"] = {"performance": perf, "qps": len(self.test_queries) / perf["avg_time"]}
                    print(f"   ✓ HNSW_full: {perf['avg_time']*1000:.2f}ms")
                
                # 5. Batch query
                def batch_search():
                    return collection.query(query_embeddings=query_vectors, n_results=10)
                
                perf = self._measure_search_time(batch_search)
                if "error" not in perf:
                    model_results["HNSW_batch"] = {"performance": perf, "qps": len(self.test_queries) / perf["avg_time"]}
                    print(f"   ✓ HNSW_batch: {perf['avg_time']*1000:.2f}ms")
                
                results[model_key] = model_results
                
            except Exception as e:
                results[model_key] = {"error": str(e)}
                print(f"   ❌ Hata: {e}")
        
        self.results["multi_model_benchmark"]["chromadb_search"] = results
        return results
    
    def benchmark_all_models_lancedb_search(self):
        """Tüm modeller için LanceDB arama benchmark - Çoklu algoritma"""
        print("\n" + "="*70)
        print("🔍 TÜM MODELLER İÇİN LANCEDB ARAMA BENCHMARK")
        print("="*70)
        
        results = {}
        
        for model_key, query_vectors in self.all_query_vectors.items():
            if not query_vectors:
                continue
            
            table_name = f"docs_{model_key}"
            db_path = os.path.join(self.db_base_path, f"lancedb/{model_key}_db")
            
            print(f"\n📊 {model_key}...")
            
            try:
                db = lancedb.connect(db_path)
                table = db.open_table(table_name)
                
                model_results = {}
                
                # 1. Default Vector Search
                def vector_search():
                    return [table.search(qv).limit(10).to_pandas() for qv in query_vectors]
                
                perf = self._measure_search_time(vector_search)
                if "error" not in perf:
                    model_results["VECTOR_default"] = {"performance": perf, "qps": len(self.test_queries) / perf["avg_time"]}
                    print(f"   ✓ VECTOR_default: {perf['avg_time']*1000:.2f}ms")
                
                # 2. Farklı limit değerleri
                for limit in [5, 20, 50, 100, 200]:
                    def limit_search(l=limit):
                        return [table.search(qv).limit(l).to_pandas() for qv in query_vectors]
                    
                    perf = self._measure_search_time(limit_search)
                    if "error" not in perf:
                        model_results[f"VECTOR_limit{limit}"] = {"performance": perf, "qps": len(self.test_queries) / perf["avg_time"]}
                        print(f"   ✓ VECTOR_limit{limit}: {perf['avg_time']*1000:.2f}ms")
                
                # 3. Select specific columns
                def select_search():
                    return [table.search(qv).limit(10).select(["text"]).to_pandas() for qv in query_vectors]
                
                perf = self._measure_search_time(select_search)
                if "error" not in perf:
                    model_results["VECTOR_select_text"] = {"performance": perf, "qps": len(self.test_queries) / perf["avg_time"]}
                    print(f"   ✓ VECTOR_select_text: {perf['avg_time']*1000:.2f}ms")
                
                # 4. Cosine metric (default)
                def cosine_search():
                    return [table.search(qv).metric("cosine").limit(10).to_pandas() for qv in query_vectors]
                
                perf = self._measure_search_time(cosine_search)
                if "error" not in perf:
                    model_results["VECTOR_cosine"] = {"performance": perf, "qps": len(self.test_queries) / perf["avg_time"]}
                    print(f"   ✓ VECTOR_cosine: {perf['avg_time']*1000:.2f}ms")
                
                # 5. L2 (Euclidean) metric
                def l2_search():
                    return [table.search(qv).metric("L2").limit(10).to_pandas() for qv in query_vectors]
                
                perf = self._measure_search_time(l2_search)
                if "error" not in perf:
                    model_results["VECTOR_L2"] = {"performance": perf, "qps": len(self.test_queries) / perf["avg_time"]}
                    print(f"   ✓ VECTOR_L2: {perf['avg_time']*1000:.2f}ms")
                
                # 6. Dot product metric
                def dot_search():
                    return [table.search(qv).metric("dot").limit(10).to_pandas() for qv in query_vectors]
                
                perf = self._measure_search_time(dot_search)
                if "error" not in perf:
                    model_results["VECTOR_dot"] = {"performance": perf, "qps": len(self.test_queries) / perf["avg_time"]}
                    print(f"   ✓ VECTOR_dot: {perf['avg_time']*1000:.2f}ms")
                
                # 7. nprobes değerleri (IVF için)
                for nprobes in [1, 8, 20, 50]:
                    def nprobes_search(np=nprobes):
                        return [table.search(qv).nprobes(np).limit(10).to_pandas() for qv in query_vectors]
                    
                    try:
                        perf = self._measure_search_time(nprobes_search)
                        if "error" not in perf:
                            model_results[f"VECTOR_nprobes{nprobes}"] = {"performance": perf, "qps": len(self.test_queries) / perf["avg_time"]}
                            print(f"   ✓ VECTOR_nprobes{nprobes}: {perf['avg_time']*1000:.2f}ms")
                    except:
                        pass
                
                # 8. Refine factor
                for refine in [1, 5, 10]:
                    def refine_search(rf=refine):
                        return [table.search(qv).refine_factor(rf).limit(10).to_pandas() for qv in query_vectors]
                    
                    try:
                        perf = self._measure_search_time(refine_search)
                        if "error" not in perf:
                            model_results[f"VECTOR_refine{refine}"] = {"performance": perf, "qps": len(self.test_queries) / perf["avg_time"]}
                            print(f"   ✓ VECTOR_refine{refine}: {perf['avg_time']*1000:.2f}ms")
                    except:
                        pass
                
                results[model_key] = model_results
                
            except Exception as e:
                results[model_key] = {"error": str(e)}
                print(f"   ❌ Hata: {e}")
        
        self.results["multi_model_benchmark"]["lancedb_search"] = results
        return results
    
    def benchmark_all_models_weaviate_search(self):
        """Tüm modeller için Weaviate arama benchmark - Çoklu algoritma"""
        print("\n" + "="*70)
        print("🔍 TÜM MODELLER İÇİN WEAVIATE ARAMA BENCHMARK")
        print("="*70)
        
        results = {}
        
        try:
            client = weaviate.connect_to_local()
        except Exception as e:
            print(f"❌ Weaviate bağlantı hatası: {e}")
            return {"error": str(e)}
        
        for model_key, query_vectors in self.all_query_vectors.items():
            if not query_vectors:
                continue
            
            collection_name = f"Docs{model_key.replace('_', '').title()}"
            print(f"\n📊 {model_key}...")
            
            try:
                collection = client.collections.get(collection_name)
                model_results = {}
                
                # 1. Near Vector Search (default HNSW)
                def near_vector_search():
                    return [collection.query.near_vector(near_vector=qv, limit=10, return_metadata=["distance"]) for qv in query_vectors]
                
                perf = self._measure_search_time(near_vector_search)
                if "error" not in perf:
                    model_results["HNSW_near_vector"] = {"performance": perf, "qps": len(self.test_queries) / perf["avg_time"]}
                    print(f"   ✓ HNSW_near_vector: {perf['avg_time']*1000:.2f}ms")
                
                # 2. Farklı limit değerleri
                for limit in [5, 20, 50, 100]:
                    def limit_search(l=limit):
                        return [collection.query.near_vector(near_vector=qv, limit=l) for qv in query_vectors]
                    
                    perf = self._measure_search_time(limit_search)
                    if "error" not in perf:
                        model_results[f"HNSW_limit{limit}"] = {"performance": perf, "qps": len(self.test_queries) / perf["avg_time"]}
                        print(f"   ✓ HNSW_limit{limit}: {perf['avg_time']*1000:.2f}ms")
                
                # 3. BM25 Search (keyword-based)
                def bm25_search():
                    return [collection.query.bm25(query=q, limit=10) for q in self.test_queries]
                
                perf = self._measure_search_time(bm25_search)
                if "error" not in perf:
                    model_results["BM25"] = {"performance": perf, "qps": len(self.test_queries) / perf["avg_time"]}
                    print(f"   ✓ BM25: {perf['avg_time']*1000:.2f}ms")
                
                # 4. Hybrid Search - farklı alpha değerleri
                for alpha in [0.0, 0.25, 0.5, 0.75, 1.0]:
                    def hybrid_search(a=alpha):
                        return [collection.query.hybrid(query=q, vector=qv, limit=10, alpha=a) for q, qv in zip(self.test_queries, query_vectors)]
                    
                    perf = self._measure_search_time(hybrid_search)
                    if "error" not in perf:
                        model_results[f"HYBRID_alpha{alpha}"] = {"performance": perf, "qps": len(self.test_queries) / perf["avg_time"]}
                        print(f"   ✓ HYBRID_alpha{alpha}: {perf['avg_time']*1000:.2f}ms")
                
                # 5. Near vector with certainty threshold
                for certainty in [0.5, 0.7, 0.9]:
                    def certainty_search(c=certainty):
                        return [collection.query.near_vector(near_vector=qv, limit=10, certainty=c) for qv in query_vectors]
                    
                    try:
                        perf = self._measure_search_time(certainty_search)
                        if "error" not in perf:
                            model_results[f"HNSW_certainty{certainty}"] = {"performance": perf, "qps": len(self.test_queries) / perf["avg_time"]}
                            print(f"   ✓ HNSW_certainty{certainty}: {perf['avg_time']*1000:.2f}ms")
                    except:
                        pass
                
                # 6. Near vector with distance threshold
                for distance in [0.3, 0.5, 0.7]:
                    def distance_search(d=distance):
                        return [collection.query.near_vector(near_vector=qv, limit=10, distance=d) for qv in query_vectors]
                    
                    try:
                        perf = self._measure_search_time(distance_search)
                        if "error" not in perf:
                            model_results[f"HNSW_distance{distance}"] = {"performance": perf, "qps": len(self.test_queries) / perf["avg_time"]}
                            print(f"   ✓ HNSW_distance{distance}: {perf['avg_time']*1000:.2f}ms")
                    except:
                        pass
                
                # 7. Fetch objects (baseline - no vector search)
                def fetch_search():
                    return [collection.query.fetch_objects(limit=10) for _ in query_vectors]
                
                perf = self._measure_search_time(fetch_search)
                if "error" not in perf:
                    model_results["FETCH_baseline"] = {"performance": perf, "qps": len(self.test_queries) / perf["avg_time"]}
                    print(f"   ✓ FETCH_baseline: {perf['avg_time']*1000:.2f}ms")
                
                # 8. BM25 with different limits
                for limit in [5, 20, 50]:
                    def bm25_limit_search(l=limit):
                        return [collection.query.bm25(query=q, limit=l) for q in self.test_queries]
                    
                    perf = self._measure_search_time(bm25_limit_search)
                    if "error" not in perf:
                        model_results[f"BM25_limit{limit}"] = {"performance": perf, "qps": len(self.test_queries) / perf["avg_time"]}
                        print(f"   ✓ BM25_limit{limit}: {perf['avg_time']*1000:.2f}ms")
                
                results[model_key] = model_results
                
            except Exception as e:
                results[model_key] = {"error": str(e)}
                print(f"   ❌ Hata: {e}")
        
        self.results["multi_model_benchmark"]["weaviate_search"] = results
        return results
    
    def _measure_search_time(self, search_func, warmup_runs=2, test_runs=5) -> Dict:
        """Arama süresini ölç"""
        for _ in range(warmup_runs):
            try:
                search_func()
            except:
                pass
        
        times = []
        for _ in range(test_runs):
            start = time.time()
            try:
                search_func()
                times.append(time.time() - start)
            except Exception as e:
                return {"error": str(e)}
        
        return {
            "avg_time": np.mean(times),
            "min_time": np.min(times),
            "max_time": np.max(times),
            "std_time": np.std(times),
            "p50_time": np.percentile(times, 50),
            "p95_time": np.percentile(times, 95),
            "p99_time": np.percentile(times, 99)
        }
    
    # ==================== MODEL VERİLERİNİ YÜKLE ====================
    def load_models_data(self) -> bool:
        """Models klasöründeki verileri yükle"""
        print("\n" + "="*70)
        print("📂 VERİLER YÜKLENİYOR")
        print("="*70)
        
        if not os.path.exists(self.models_path):
            print(f"❌ Models klasörü bulunamadı: {self.models_path}")
            return False
        
        json_files = glob.glob(os.path.join(self.models_path, "*.json"))
        txt_files = glob.glob(os.path.join(self.models_path, "*.txt"))
        
        documents = []
        
        for json_file in json_files:
            try:
                with open(json_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    if isinstance(data, list):
                        for item in data:
                            if isinstance(item, dict):
                                text = item.get('text', item.get('content', str(item)))
                                documents.append({"text": text, "source": os.path.basename(json_file)})
                            else:
                                documents.append({"text": str(item), "source": os.path.basename(json_file)})
                    elif isinstance(data, dict):
                        text = data.get('text', data.get('content', str(data)))
                        documents.append({"text": text, "source": os.path.basename(json_file)})
            except:
                pass
        
        for txt_file in txt_files:
            try:
                with open(txt_file, 'r', encoding='utf-8') as f:
                    content = f.read()
                    chunks = [content[i:i+500] for i in range(0, len(content), 500)]
                    for chunk in chunks:
                        if chunk.strip():
                            documents.append({"text": chunk.strip(), "source": os.path.basename(txt_file)})
            except:
                pass
        
        if not documents:
            documents = self._generate_sample_data()
        
        self.documents = documents
        print(f"✅ Toplam {len(self.documents)} doküman yüklendi")
        return True
    
    def _generate_sample_data(self) -> List[Dict]:
        """Örnek veri oluştur"""
        sample_texts = [
            "Artificial intelligence is transforming healthcare with advanced diagnostic tools.",
            "Machine learning algorithms can predict patient outcomes with high accuracy.",
            "Deep learning models are revolutionizing medical image analysis.",
            "Natural language processing enables better understanding of clinical notes.",
            "Computer vision systems can detect diseases from X-rays and MRI scans.",
            "Reinforcement learning is being applied to optimize treatment plans.",
            "Transformer models have achieved state-of-the-art results in NLP tasks.",
            "Convolutional neural networks excel at image classification tasks.",
            "Recurrent neural networks are effective for sequence modeling.",
            "Generative adversarial networks can create synthetic medical images."
        ] * 100  # 1000 doküman
        
        return [{"text": text, "source": "sample"} for text in sample_texts]
    
    # ==================== SONUÇLARI KAYDET ====================
    def save_comprehensive_results(self):
        """Kapsamlı sonuçları kaydet"""
        output_dir = self.base_path
        excel_file = os.path.join(output_dir, "multi_model_benchmark_results.xlsx")
        json_file = os.path.join(output_dir, "multi_model_benchmark_results.json")
        
        # JSON kaydet
        with open(json_file, 'w', encoding='utf-8') as f:
            json.dump(self.results, f, indent=2, ensure_ascii=False, default=str)
        print(f"\n💾 JSON: {json_file}")
        
        # Excel oluştur
        wb = openpyxl.Workbook()
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        gold_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        silver_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
        bronze_fill = PatternFill(start_color="CD7F32", end_color="CD7F32", fill_type="solid")
        green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        center = Alignment(horizontal='center', vertical='center')
        left = Alignment(horizontal='left', vertical='center')
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        # ==================== 1. GENEL ÖZET SAYFASI ====================
        ws_summary = wb.active
        ws_summary.title = "Genel Özet"
        
        ws_summary.merge_cells('A1:F1')
        ws_summary['A1'] = "BENCHMARK GENEL ÖZETİ"
        ws_summary['A1'].font = Font(bold=True, size=16)
        ws_summary['A1'].alignment = center
        
        # Meta bilgiler
        ws_summary['A3'] = "Test Tarihi:"
        ws_summary['B3'] = self.results["metadata"]["date"]
        ws_summary['A4'] = "Sorgu Sayısı:"
        ws_summary['B4'] = self.results["metadata"]["query_count"]
        ws_summary['A5'] = "Doküman Sayısı:"
        ws_summary['B5'] = len(self.documents)
        
        # Yüklenen modeller
        ws_summary['A7'] = "YÜKLENEN MODELLER:"
        ws_summary['A7'].font = Font(bold=True)
        row = 8
        for model_key, model_data in self.model_info.items():
            status = model_data.get('status', 'N/A')
            dim = model_data.get('vector_dim', 'N/A')
            ws_summary[f'A{row}'] = f"  • {model_data['name']}"
            ws_summary[f'B{row}'] = f"Durum: {status}"
            ws_summary[f'C{row}'] = f"Dim: {dim}"
            row += 1
        
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws_summary.column_dimensions[col].width = 25
        
        # ==================== 2. YAZMA BENCHMARK SAYFASI ====================
        ws_write = wb.create_sheet("Yazma Benchmark")
        ws_write.merge_cells('A1:E1')
        ws_write['A1'] = "TÜM MODELLER - YAZMA BENCHMARK SONUÇLARI"
        ws_write['A1'].font = Font(bold=True, size=14)
        ws_write['A1'].alignment = center
        
        databases = ['milvus_write', 'qdrant_write', 'chromadb_write', 'lancedb_write', 'weaviate_write']
        row = 3
        
        for db in databases:
            db_data = self.results.get("multi_model_benchmark", {}).get(db, {})
            if not db_data or "error" in db_data:
                continue
            
            ws_write[f'A{row}'] = db.replace('_write', '').upper()
            ws_write[f'A{row}'].font = Font(bold=True, size=12)
            ws_write[f'A{row}'].fill = green_fill
            row += 1
            
            headers = ["Model", "Süre (s)", "Kayıt Sayısı", "Vektör Dim", "Kayıt/sn"]
            for col, h in enumerate(headers, 1):
                cell = ws_write.cell(row=row, column=col, value=h)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center
                cell.border = border
            row += 1
            
            for model_key, model_data in db_data.items():
                if isinstance(model_data, dict) and model_data.get("status") == "success":
                    ws_write.cell(row=row, column=1, value=model_key).border = border
                    ws_write.cell(row=row, column=2, value=round(model_data["write_time"], 4)).border = border
                    ws_write.cell(row=row, column=3, value=model_data["record_count"]).border = border
                    ws_write.cell(row=row, column=4, value=model_data["vector_dim"]).border = border
                    ws_write.cell(row=row, column=5, value=round(model_data["record_count"] / model_data["write_time"], 2)).border = border
                    for col in range(1, 6):
                        ws_write.cell(row=row, column=col).alignment = center
                    row += 1
            row += 2
        
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws_write.column_dimensions[col].width = 20
        
        # ==================== 3. TÜM ARAMA SONUÇLARI SAYFASI ====================
        ws_all_search = wb.create_sheet("Tüm Arama Sonuçları")
        ws_all_search.merge_cells('A1:I1')
        ws_all_search['A1'] = "TÜM ARAMA ALGORİTMALARI - DETAYLI SONUÇLAR"
        ws_all_search['A1'].font = Font(bold=True, size=14)
        ws_all_search['A1'].alignment = center
        
        # Tüm arama sonuçlarını topla
        all_results = []
        search_dbs = ['milvus_search', 'qdrant_search', 'chromadb_search', 'lancedb_search', 'weaviate_search']
        
        for db in search_dbs:
            db_data = self.results.get("multi_model_benchmark", {}).get(db, {})
            if not db_data:
                continue
            
            for model_key, model_data in db_data.items():
                if isinstance(model_data, dict) and "error" not in model_data:
                    for algo, algo_data in model_data.items():
                        if isinstance(algo_data, dict) and "performance" in algo_data:
                            perf = algo_data["performance"]
                            all_results.append({
                                "database": db.replace('_search', ''),
                                "model": model_key,
                                "algorithm": algo,
                                "avg_ms": perf["avg_time"] * 1000,
                                "min_ms": perf["min_time"] * 1000,
                                "max_ms": perf["max_time"] * 1000,
                                "std_ms": perf.get("std_time", 0) * 1000,
                                "p50_ms": perf.get("p50_time", 0) * 1000,
                                "p95_ms": perf.get("p95_time", 0) * 1000,
                                "p99_ms": perf.get("p99_time", 0) * 1000,
                                "qps": algo_data.get("qps", 0)
                            })
        
        all_results.sort(key=lambda x: x["avg_ms"])
        
        headers = ["Sıra", "Veritabanı", "Model", "Algoritma", "Ort (ms)", "Min (ms)", "Max (ms)", "Std (ms)", "P95 (ms)", "QPS"]
        row = 3
        for col, h in enumerate(headers, 1):
            cell = ws_all_search.cell(row=row, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border
        row += 1
        
        # TÜM sonuçları yaz (sadece 50 değil)
        for i, r in enumerate(all_results, 1):
            ws_all_search.cell(row=row, column=1, value=i).border = border
            ws_all_search.cell(row=row, column=2, value=r["database"]).border = border
            ws_all_search.cell(row=row, column=3, value=r["model"]).border = border
            ws_all_search.cell(row=row, column=4, value=r["algorithm"]).border = border
            ws_all_search.cell(row=row, column=5, value=round(r["avg_ms"], 4)).border = border
            ws_all_search.cell(row=row, column=6, value=round(r["min_ms"], 4)).border = border
            ws_all_search.cell(row=row, column=7, value=round(r["max_ms"], 4)).border = border
            ws_all_search.cell(row=row, column=8, value=round(r["std_ms"], 4)).border = border
            ws_all_search.cell(row=row, column=9, value=round(r["p95_ms"], 4)).border = border
            ws_all_search.cell(row=row, column=10, value=round(r["qps"], 2)).border = border
            
            for col in range(1, 11):
                cell = ws_all_search.cell(row=row, column=col)
                cell.alignment = center
                if i == 1: cell.fill = gold_fill
                elif i == 2: cell.fill = silver_fill
                elif i == 3: cell.fill = bronze_fill
            row += 1
        
        ws_all_search.column_dimensions['A'].width = 8
        ws_all_search.column_dimensions['B'].width = 12
        ws_all_search.column_dimensions['C'].width = 22
        ws_all_search.column_dimensions['D'].width = 22
        for col in ['E', 'F', 'G', 'H', 'I', 'J']:
            ws_all_search.column_dimensions[col].width = 12
        
        # ==================== 4. VERİTABANI BAZINDA ARAMA ====================
        for db in search_dbs:
            db_name = db.replace('_search', '')
            ws_db = wb.create_sheet(f"{db_name.upper()} Arama")
            
            ws_db.merge_cells('A1:H1')
            ws_db['A1'] = f"{db_name.upper()} - ARAMA ALGORİTMALARI DETAY"
            ws_db['A1'].font = Font(bold=True, size=14)
            ws_db['A1'].alignment = center
            
            db_data = self.results.get("multi_model_benchmark", {}).get(db, {})
            if not db_data:
                ws_db['A3'] = "Veri bulunamadı"
                continue
            
            row = 3
            for model_key, model_data in db_data.items():
                if isinstance(model_data, dict) and "error" not in model_data:
                    ws_db[f'A{row}'] = f"Model: {model_key}"
                    ws_db[f'A{row}'].font = Font(bold=True, size=11)
                    ws_db[f'A{row}'].fill = green_fill
                    row += 1
                    
                    headers = ["Algoritma", "Ort (ms)", "Min (ms)", "Max (ms)", "Std (ms)", "P50 (ms)", "P95 (ms)", "QPS"]
                    for col, h in enumerate(headers, 1):
                        cell = ws_db.cell(row=row, column=col, value=h)
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = center
                        cell.border = border
                    row += 1
                    
                    # Algoritmaları süreye göre sırala
                    algo_list = []
                    for algo, algo_data in model_data.items():
                        if isinstance(algo_data, dict) and "performance" in algo_data:
                            algo_list.append((algo, algo_data))
                    algo_list.sort(key=lambda x: x[1]["performance"]["avg_time"])
                    
                    for algo, algo_data in algo_list:
                        perf = algo_data["performance"]
                        ws_db.cell(row=row, column=1, value=algo).border = border
                        ws_db.cell(row=row, column=2, value=round(perf["avg_time"]*1000, 4)).border = border
                        ws_db.cell(row=row, column=3, value=round(perf["min_time"]*1000, 4)).border = border
                        ws_db.cell(row=row, column=4, value=round(perf["max_time"]*1000, 4)).border = border
                        ws_db.cell(row=row, column=5, value=round(perf.get("std_time", 0)*1000, 4)).border = border
                        ws_db.cell(row=row, column=6, value=round(perf.get("p50_time", 0)*1000, 4)).border = border
                        ws_db.cell(row=row, column=7, value=round(perf.get("p95_time", 0)*1000, 4)).border = border
                        ws_db.cell(row=row, column=8, value=round(algo_data.get("qps", 0), 2)).border = border
                        for col in range(1, 9):
                            ws_db.cell(row=row, column=col).alignment = center
                        row += 1
                    row += 2
            
            ws_db.column_dimensions['A'].width = 25
            for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
                ws_db.column_dimensions[col].width = 12
        
        # ==================== 5. MODEL BAZINDA KARŞILAŞTIRMA ====================
        ws_model = wb.create_sheet("Model Karşılaştırma")
        ws_model.merge_cells('A1:G1')
        ws_model['A1'] = "MODEL BAZINDA EN İYİ SONUÇLAR"
        ws_model['A1'].font = Font(bold=True, size=14)
        ws_model['A1'].alignment = center
        
        # Her model için en iyi sonuçları bul
        model_best = {}
        for r in all_results:
            model = r["model"]
            if model not in model_best or r["avg_ms"] < model_best[model]["avg_ms"]:
                model_best[model] = r
        
        headers = ["Model", "En İyi DB", "En İyi Algoritma", "Süre (ms)", "QPS", "Toplam Test"]
        row = 3
        for col, h in enumerate(headers, 1):
            cell = ws_model.cell(row=row, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border
        row += 1
        
        for model, best in sorted(model_best.items(), key=lambda x: x[1]["avg_ms"]):
            test_count = len([r for r in all_results if r["model"] == model])
            ws_model.cell(row=row, column=1, value=model).border = border
            ws_model.cell(row=row, column=2, value=best["database"]).border = border
            ws_model.cell(row=row, column=3, value=best["algorithm"]).border = border
            ws_model.cell(row=row, column=4, value=round(best["avg_ms"], 4)).border = border
            ws_model.cell(row=row, column=5, value=round(best["qps"], 2)).border = border
            ws_model.cell(row=row, column=6, value=test_count).border = border
            for col in range(1, 7):
                ws_model.cell(row=row, column=col).alignment = center
            row += 1
        
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws_model.column_dimensions[col].width = 22
        
        # ==================== 6. ALGORİTMA KATEGORİLERİ ====================
        ws_algo = wb.create_sheet("Algoritma Kategorileri")
        ws_algo.merge_cells('A1:F1')
        ws_algo['A1'] = "ALGORİTMA KATEGORİLERİ PERFORMANS ANALİZİ"
        ws_algo['A1'].font = Font(bold=True, size=14)
        ws_algo['A1'].alignment = center
        
        categories = {
            "HNSW Tabanlı": ["HNSW", "near_vector"],
            "Batch İşlem": ["batch"],
            "Limit Varyasyonları": ["limit"],
            "BM25/Keyword": ["BM25", "TEXT"],
            "Hybrid": ["HYBRID"],
            "Exact/Brute Force": ["EXACT", "bruteforce"],
            "Metric Varyasyonları": ["cosine", "L2", "dot"],
            "Quantization": ["quantized"],
            "Parametre Testi": ["ef", "nprobe", "refine", "nprobes"]
        }
        
        headers = ["Kategori", "Test Sayısı", "Ort Süre (ms)", "Min Süre (ms)", "En İyi Kombinasyon"]
        row = 3
        for col, h in enumerate(headers, 1):
            cell = ws_algo.cell(row=row, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border
        row += 1
        
        for category, patterns in categories.items():
            category_results = []
            for r in all_results:
                for pattern in patterns:
                    if pattern.lower() in r["algorithm"].lower():
                        category_results.append(r)
                        break
            
            if category_results:
                avg_time = np.mean([r["avg_ms"] for r in category_results])
                min_time = min([r["avg_ms"] for r in category_results])
                best = min(category_results, key=lambda x: x["avg_ms"])
                best_combo = f"{best['database']}/{best['model']}/{best['algorithm']}"
                
                ws_algo.cell(row=row, column=1, value=category).border = border
                ws_algo.cell(row=row, column=2, value=len(category_results)).border = border
                ws_algo.cell(row=row, column=3, value=round(avg_time, 4)).border = border
                ws_algo.cell(row=row, column=4, value=round(min_time, 4)).border = border
                ws_algo.cell(row=row, column=5, value=best_combo).border = border
                for col in range(1, 6):
                    ws_algo.cell(row=row, column=col).alignment = center if col < 5 else left
                row += 1
        
        ws_algo.column_dimensions['A'].width = 22
        ws_algo.column_dimensions['B'].width = 12
        ws_algo.column_dimensions['C'].width = 15
        ws_algo.column_dimensions['D'].width = 15
        ws_algo.column_dimensions['E'].width = 50
        
        # ==================== 7. VERİTABANI KARŞILAŞTIRMA ====================
        ws_db_compare = wb.create_sheet("Veritabanı Karşılaştırma")
        ws_db_compare.merge_cells('A1:H1')
        ws_db_compare['A1'] = "VERİTABANI PERFORMANS KARŞILAŞTIRMASI"
        ws_db_compare['A1'].font = Font(bold=True, size=14)
        ws_db_compare['A1'].alignment = center
        
        headers = ["Veritabanı", "Toplam Test", "Ort Süre (ms)", "Min Süre (ms)", "Max Süre (ms)", "Ort QPS", "En İyi Algoritma"]
        row = 3
        for col, h in enumerate(headers, 1):
            cell = ws_db_compare.cell(row=row, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border
        row += 1
        
        db_stats = {}
        for r in all_results:
            db = r["database"]
            if db not in db_stats:
                db_stats[db] = {"results": [], "best": None}
            db_stats[db]["results"].append(r)
            if db_stats[db]["best"] is None or r["avg_ms"] < db_stats[db]["best"]["avg_ms"]:
                db_stats[db]["best"] = r
        
        for db, stats in sorted(db_stats.items(), key=lambda x: np.mean([r["avg_ms"] for r in x[1]["results"]])):
            results = stats["results"]
            avg_time = np.mean([r["avg_ms"] for r in results])
            min_time = min([r["avg_ms"] for r in results])
            max_time = max([r["avg_ms"] for r in results])
            avg_qps = np.mean([r["qps"] for r in results])
            best_algo = stats["best"]["algorithm"]
            
            ws_db_compare.cell(row=row, column=1, value=db).border = border
            ws_db_compare.cell(row=row, column=2, value=len(results)).border = border
            ws_db_compare.cell(row=row, column=3, value=round(avg_time, 4)).border = border
            ws_db_compare.cell(row=row, column=4, value=round(min_time, 4)).border = border
            ws_db_compare.cell(row=row, column=5, value=round(max_time, 4)).border = border
            ws_db_compare.cell(row=row, column=6, value=round(avg_qps, 2)).border = border
            ws_db_compare.cell(row=row, column=7, value=best_algo).border = border
            for col in range(1, 8):
                ws_db_compare.cell(row=row, column=col).alignment = center
            row += 1
        
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            ws_db_compare.column_dimensions[col].width = 18
        
        # Dosyayı kaydet
        wb.save(excel_file)
        print(f"💾 Excel: {excel_file}")
        print(f"   📄 Sayfalar:")
        print(f"      - Genel Özet: Meta bilgiler ve yüklenen modeller")
        print(f"      - Yazma Benchmark: Tüm yazma işlemleri")
        print(f"      - Tüm Arama Sonuçları: {len(all_results)} arama testi")
        print(f"      - MILVUS/QDRANT/CHROMADB/LANCEDB/WEAVIATE Arama: Veritabanı detayları")
        print(f"      - Model Karşılaştırma: Model bazında en iyi sonuçlar")
        print(f"      - Algoritma Kategorileri: Kategori bazında analiz")
        print(f"      - Veritabanı Karşılaştırma: DB performans özeti")

    def print_comprehensive_summary(self):
        """Kapsamlı özet yazdır"""
        print("\n" + "="*80)
        print("📊 KAPSAMLI BENCHMARK SONUÇ ÖZETİ")
        print("="*80)
        
        # Yazma sonuçları
        print("\n📝 YAZMA BENCHMARK (Model x Veritabanı):")
        print("-" * 60)
        
        for db in ['milvus_write', 'qdrant_write', 'chromadb_write', 'lancedb_write', 'weaviate_write']:
            db_data = self.results.get("multi_model_benchmark", {}).get(db, {})
            if db_data and "error" not in db_data:
                print(f"\n  {db.replace('_write', '').upper()}:")
                for model, data in db_data.items():
                    if isinstance(data, dict) and data.get("status") == "success":
                        print(f"    • {model}: {data['write_time']:.2f}s")
        
        # Arama sonuçları
        print("\n🔍 ARAMA BENCHMARK - EN HIZLI 20:")
        print("-" * 80)
        
        all_results = []
        for db in ['milvus_search', 'qdrant_search', 'chromadb_search', 'lancedb_search', 'weaviate_search']:
            db_data = self.results.get("multi_model_benchmark", {}).get(db, {})
            if db_data:
                for model, model_data in db_data.items():
                    if isinstance(model_data, dict) and "error" not in model_data:
                        for algo, algo_data in model_data.items():
                            if isinstance(algo_data, dict) and "performance" in algo_data:
                                all_results.append((
                                    db.replace('_search', ''),
                                    model,
                                    algo,
                                    algo_data["performance"]["avg_time"] * 1000,
                                    algo_data.get("qps", 0)
                                ))
        
        all_results.sort(key=lambda x: x[3])
        
        for i, (db, model, algo, time_ms, qps) in enumerate(all_results[:20], 1):
            emoji = "🥇" if i == 1 else "🥈" if i == 2 else "🥉" if i == 3 else f"{i:2}."
            print(f"  {emoji} {db:<10} | {model:<22} | {algo:<20} | {time_ms:8.2f}ms | QPS: {qps:8.0f}")
        
        print(f"\n📊 Toplam {len(all_results)} arama testi yapıldı.")
    
    def run_full_multi_model_benchmark(self):
        """Tüm modeller ve veritabanları için kapsamlı benchmark"""
        print("\n" + "🚀"*40)
        print("    KAPSAMLI MULTI-MODEL BENCHMARK BAŞLIYOR")
        print("🚀"*40)
        print(f"\n📅 Tarih: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        
        # 1. Verileri yükle
        if not self.load_models_data():
            print("❌ Veri yükleme başarısız!")
            return
        
        # 2. Tüm modeller için embedding hesapla
        self.prepare_all_embeddings()
        
        # 3. Tüm veritabanlarına yaz
        print("\n" + "="*70)
        print("📝 TÜM MODELLER İÇİN YAZMA BAŞLIYOR")
        print("="*70)
        
        self.write_all_models_to_milvus()
        self.write_all_models_to_qdrant()
        self.write_all_models_to_chromadb()
        self.write_all_models_to_lancedb()
        self.write_all_models_to_weaviate()
        
        # 4. Tüm veritabanlarında arama benchmark
        print("\n" + "="*70)
        print("🔍 TÜM MODELLER İÇİN ARAMA BENCHMARK BAŞLIYOR")
        print("="*70)
        
        self.benchmark_all_models_milvus_search()
        self.benchmark_all_models_qdrant_search()
        self.benchmark_all_models_chromadb_search()
        self.benchmark_all_models_lancedb_search()
        self.benchmark_all_models_weaviate_search()
        
        # 5. Sonuçları kaydet
        self.print_comprehensive_summary()
        self.save_comprehensive_results()
        
        print("\n" + "✅"*40)
        print("    KAPSAMLI BENCHMARK TAMAMLANDI!")
        print("✅"*40)


if __name__ == "__main__":
    print("="*70)
    print("🚀 VECTOR DATABASE MULTI-MODEL BENCHMARK")
    print("="*70)
    
    try:
        # Benchmark nesnesini oluştur
        benchmark = VectorDatabaseBenchmark()
        
        # Kapsamlı benchmark'ı çalıştır
        benchmark.run_full_multi_model_benchmark()
        
    except KeyboardInterrupt:
        print("\n\n⚠️ Kullanıcı tarafından iptal edildi.")
    except Exception as e:
        print(f"\n❌ Hata oluştu: {e}")
        import traceback
        traceback.print_exc()