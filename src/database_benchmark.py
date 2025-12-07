import time
import psutil
import os
import json
from typing import Dict, List
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import lancedb
import chromadb
from qdrant_client import QdrantClient
from pymilvus import MilvusClient
import weaviate
import psycopg2
from sentence_transformers import SentenceTransformer

class DatabaseBenchmark:
    def __init__(self):
        self.results = {
            "lancedb": {},
            "chromadb": {},
            "qdrant": {},
            "milvus": {},
            "weaviate": {},
            "pgvector": {}
        }
        self.model = SentenceTransformer("all-MiniLM-L6-v2")
        
        # Test sorguları
        self.test_queries = [
            "artificial intelligence healthcare",
            "machine learning medical diagnosis",
            "deep learning neural networks",
            "natural language processing",
            "computer vision medical imaging"
        ]
        
    def measure_memory(self) -> float:
        """Bellek kullanımını MB cinsinden ölç"""
        process = psutil.Process(os.getpid())
        return process.memory_info().rss / 1024 / 1024
    
    def measure_disk_size(self, path: str) -> float:
        """Klasör boyutunu MB cinsinden ölç"""
        if not os.path.exists(path):
            return 0
        total_size = 0
        for dirpath, dirnames, filenames in os.walk(path):
            for filename in filenames:
                filepath = os.path.join(dirpath, filename)
                try:
                    total_size += os.path.getsize(filepath)
                except:
                    pass
        return total_size / 1024 / 1024
    
    # ==================== LanceDB Benchmark ====================
    def benchmark_lancedb(self):
        """LanceDB performansını ölç"""
        print("\n" + "="*60)
        print("📊 LanceDB BENCHMARK")
        print("="*60)
        
        try:
            db_path = "/home/ugo/Documents/Python/bitirememe projesi/DB/lanceDatabase/db"
            
            if not os.path.exists(db_path):
                print(f"⚠ LanceDB veritabanı bulunamadı: {db_path}")
                self.results["lancedb"]["error"] = "Database not found"
                return
            
            # Bağlantı zamanı
            start = time.time()
            db = lancedb.connect(db_path)
            connection_time = time.time() - start
            self.results["lancedb"]["connection_time"] = connection_time
            print(f"✓ Bağlantı zamanı: {connection_time:.4f}s")
            
            # Tablo açma zamanı
            start = time.time()
            table = db.open_table("documents")
            open_time = time.time() - start
            self.results["lancedb"]["open_table_time"] = open_time
            print(f"✓ Tablo açma zamanı: {open_time:.4f}s")
            
            # Kayıt sayısı
            start = time.time()
            count = len(table.to_pandas())
            count_time = time.time() - start
            self.results["lancedb"]["record_count"] = count
            self.results["lancedb"]["count_time"] = count_time
            print(f"✓ Toplam kayıt: {count} ({count_time:.4f}s)")
            
            # Arama performansı (soğuk başlangıç)
            print("\n🔍 Soğuk başlangıç aramaları...")
            cold_search_times = []
            for i, query in enumerate(self.test_queries, 1):
                start = time.time()
                results = table.search(query).limit(5).to_pandas()
                search_time = time.time() - start
                cold_search_times.append(search_time)
                print(f"  Sorgu {i}: {search_time:.4f}s")
            
            avg_cold_search = sum(cold_search_times) / len(cold_search_times)
            self.results["lancedb"]["avg_cold_search_time"] = avg_cold_search
            self.results["lancedb"]["min_cold_search_time"] = min(cold_search_times)
            self.results["lancedb"]["max_cold_search_time"] = max(cold_search_times)
            print(f"  Ortalama: {avg_cold_search:.4f}s")
            
            # Arama performansı (sıcak - önbellekli)
            print("\n🔥 Sıcak (cached) aramalar...")
            hot_search_times = []
            for i, query in enumerate(self.test_queries, 1):
                start = time.time()
                results = table.search(query).limit(5).to_pandas()
                search_time = time.time() - start
                hot_search_times.append(search_time)
                print(f"  Sorgu {i}: {search_time:.4f}s")
            
            avg_hot_search = sum(hot_search_times) / len(hot_search_times)
            self.results["lancedb"]["avg_hot_search_time"] = avg_hot_search
            self.results["lancedb"]["min_hot_search_time"] = min(hot_search_times)
            self.results["lancedb"]["max_hot_search_time"] = max(hot_search_times)
            print(f"  Ortalama: {avg_hot_search:.4f}s")
            
            # Disk boyutu
            disk_size = self.measure_disk_size(db_path)
            self.results["lancedb"]["disk_size_mb"] = disk_size
            print(f"\n💾 Disk boyutu: {disk_size:.2f} MB")
            
            # Bellek kullanımı
            start_mem = self.measure_memory()
            _ = table.to_pandas()
            end_mem = self.measure_memory()
            memory_used = end_mem - start_mem
            self.results["lancedb"]["memory_used_mb"] = memory_used
            print(f"🧠 Bellek kullanımı: {memory_used:.2f} MB")
            
            print("✅ LanceDB benchmark tamamlandı")
            
        except Exception as e:
            print(f"❌ Hata: {e}")
            self.results["lancedb"]["error"] = str(e)
    
    # ==================== ChromaDB Benchmark ====================
    def benchmark_chromadb(self):
        """ChromaDB performansını ölç"""
        print("\n" + "="*60)
        print("📊 ChromaDB BENCHMARK")
        print("="*60)
        
        try:
            db_path = "/home/ugo/Documents/Python/bitirememe projesi/DB/chorame/yerel_veritabani"
            
            if not os.path.exists(db_path):
                print(f"⚠ ChromaDB veritabanı bulunamadı: {db_path}")
                self.results["chromadb"]["error"] = "Database not found"
                return
            
            # Bağlantı zamanı
            start = time.time()
            client = chromadb.PersistentClient(path=db_path)
            connection_time = time.time() - start
            self.results["chromadb"]["connection_time"] = connection_time
            print(f"✓ Bağlantı zamanı: {connection_time:.4f}s")
            
            # Collection açma
            start = time.time()
            collection = client.get_or_create_collection(name="dokumanlarim")
            collection_time = time.time() - start
            self.results["chromadb"]["collection_open_time"] = collection_time
            print(f"✓ Collection açma zamanı: {collection_time:.4f}s")
            
            # Kayıt sayısı
            start = time.time()
            count = collection.count()
            count_time = time.time() - start
            self.results["chromadb"]["record_count"] = count
            self.results["chromadb"]["count_time"] = count_time
            print(f"✓ Toplam kayıt: {count} ({count_time:.4f}s)")
            
            # Soğuk arama
            print("\n🔍 Soğuk başlangıç aramaları...")
            cold_search_times = []
            for i, query in enumerate(self.test_queries, 1):
                start = time.time()
                results = collection.query(query_texts=[query], n_results=5)
                search_time = time.time() - start
                cold_search_times.append(search_time)
                print(f"  Sorgu {i}: {search_time:.4f}s")
            
            avg_cold_search = sum(cold_search_times) / len(cold_search_times)
            self.results["chromadb"]["avg_cold_search_time"] = avg_cold_search
            self.results["chromadb"]["min_cold_search_time"] = min(cold_search_times)
            self.results["chromadb"]["max_cold_search_time"] = max(cold_search_times)
            print(f"  Ortalama: {avg_cold_search:.4f}s")
            
            # Sıcak arama
            print("\n🔥 Sıcak (cached) aramalar...")
            hot_search_times = []
            for i, query in enumerate(self.test_queries, 1):
                start = time.time()
                results = collection.query(query_texts=[query], n_results=5)
                search_time = time.time() - start
                hot_search_times.append(search_time)
                print(f"  Sorgu {i}: {search_time:.4f}s")
            
            avg_hot_search = sum(hot_search_times) / len(hot_search_times)
            self.results["chromadb"]["avg_hot_search_time"] = avg_hot_search
            self.results["chromadb"]["min_hot_search_time"] = min(hot_search_times)
            self.results["chromadb"]["max_hot_search_time"] = max(hot_search_times)
            print(f"  Ortalama: {avg_hot_search:.4f}s")
            
            # Disk boyutu
            disk_size = self.measure_disk_size(db_path)
            self.results["chromadb"]["disk_size_mb"] = disk_size
            print(f"\n💾 Disk boyutu: {disk_size:.2f} MB")
            
            print("✅ ChromaDB benchmark tamamlandı")
            
        except Exception as e:
            print(f"❌ Hata: {e}")
            self.results["chromadb"]["error"] = str(e)
    
    # ==================== Qdrant Benchmark ====================
    def benchmark_qdrant(self):
        """Qdrant performansını ölç"""
        print("\n" + "="*60)
        print("📊 Qdrant BENCHMARK")
        print("="*60)
        
        try:
            # Bağlantı zamanı
            start = time.time()
            client = QdrantClient(host="localhost", port=6333)
            connection_time = time.time() - start
            self.results["qdrant"]["connection_time"] = connection_time
            print(f"✓ Bağlantı zamanı: {connection_time:.4f}s")
            
            # Collection bilgisi
            try:
                start = time.time()
                collection_info = client.get_collection("test_collection")
                info_time = time.time() - start
                self.results["qdrant"]["info_time"] = info_time
                print(f"✓ Collection info zamanı: {info_time:.4f}s")
                
                # Kayıt sayısı
                count = collection_info.points_count
                self.results["qdrant"]["record_count"] = count
                print(f"✓ Toplam kayıt: {count}")
                
                # Soğuk arama
                print("\n🔍 Soğuk başlangıç aramaları...")
                cold_search_times = []
                for i, query in enumerate(self.test_queries, 1):
                    query_vector = self.model.encode(query).tolist()
                    start = time.time()
                    results = client.query_points(
                        collection_name="test_collection",
                        query=query_vector,
                        limit=5,
                        with_payload=True
                    )
                    search_time = time.time() - start
                    cold_search_times.append(search_time)
                    print(f"  Sorgu {i}: {search_time:.4f}s")
                
                avg_cold_search = sum(cold_search_times) / len(cold_search_times)
                self.results["qdrant"]["avg_cold_search_time"] = avg_cold_search
                self.results["qdrant"]["min_cold_search_time"] = min(cold_search_times)
                self.results["qdrant"]["max_cold_search_time"] = max(cold_search_times)
                print(f"  Ortalama: {avg_cold_search:.4f}s")
                
                # Sıcak arama
                print("\n🔥 Sıcak (cached) aramalar...")
                hot_search_times = []
                for i, query in enumerate(self.test_queries, 1):
                    query_vector = self.model.encode(query).tolist()
                    start = time.time()
                    results = client.query_points(
                        collection_name="test_collection",
                        query=query_vector,
                        limit=5,
                        with_payload=True
                    )
                    search_time = time.time() - start
                    hot_search_times.append(search_time)
                    print(f"  Sorgu {i}: {search_time:.4f}s")
                
                avg_hot_search = sum(hot_search_times) / len(hot_search_times)
                self.results["qdrant"]["avg_hot_search_time"] = avg_hot_search
                self.results["qdrant"]["min_hot_search_time"] = min(hot_search_times)
                self.results["qdrant"]["max_hot_search_time"] = max(hot_search_times)
                print(f"  Ortalama: {avg_hot_search:.4f}s")
                
                print("✅ Qdrant benchmark tamamlandı")
                
            except Exception as e:
                print(f"⚠ Qdrant collection bulunamadı veya boş: {e}")
                self.results["qdrant"]["error"] = "Collection not found or empty"
            
        except Exception as e:
            print(f"❌ Qdrant bağlantı hatası (sunucu çalışmıyor olabilir): {e}")
            self.results["qdrant"]["error"] = f"Connection failed: {str(e)}"
    
    # ==================== Milvus Benchmark ====================
    def benchmark_milvus(self):
        """Milvus performansını ölç"""
        print("\n" + "="*60)
        print("📊 Milvus BENCHMARK")
        print("="*60)
        
        try:
            db_path = "/home/ugo/Documents/Python/bitirememe projesi/DB/milvus/milvus_demo.db"
            
            if not os.path.exists(db_path):
                print(f"⚠ Milvus veritabanı bulunamadı: {db_path}")
                self.results["milvus"]["error"] = "Database not found"
                return
            
            # Bağlantı zamanı
            start = time.time()
            client = MilvusClient(db_path)
            connection_time = time.time() - start
            self.results["milvus"]["connection_time"] = connection_time
            print(f"✓ Bağlantı zamanı: {connection_time:.4f}s")
            
            # Kayıt sayısı
            start = time.time()
            stats = client.get_collection_stats(collection_name="documents")
            count = stats['row_count']
            count_time = time.time() - start
            self.results["milvus"]["record_count"] = count
            self.results["milvus"]["count_time"] = count_time
            print(f"✓ Toplam kayıt: {count} ({count_time:.4f}s)")
            
            # Soğuk arama
            print("\n🔍 Soğuk başlangıç aramaları...")
            cold_search_times = []
            for i, query in enumerate(self.test_queries, 1):
                query_vector = self.model.encode(query).tolist()
                start = time.time()
                results = client.search(
                    collection_name="documents",
                    data=[query_vector],
                    limit=5,
                    output_fields=["metin", "chunk_id", "doc_id"]
                )
                search_time = time.time() - start
                cold_search_times.append(search_time)
                print(f"  Sorgu {i}: {search_time:.4f}s")
            
            avg_cold_search = sum(cold_search_times) / len(cold_search_times)
            self.results["milvus"]["avg_cold_search_time"] = avg_cold_search
            self.results["milvus"]["min_cold_search_time"] = min(cold_search_times)
            self.results["milvus"]["max_cold_search_time"] = max(cold_search_times)
            print(f"  Ortalama: {avg_cold_search:.4f}s")
            
            # Sıcak arama
            print("\n🔥 Sıcak (cached) aramalar...")
            hot_search_times = []
            for i, query in enumerate(self.test_queries, 1):
                query_vector = self.model.encode(query).tolist()
                start = time.time()
                results = client.search(
                    collection_name="documents",
                    data=[query_vector],
                    limit=5,
                    output_fields=["metin", "chunk_id", "doc_id"]
                )
                search_time = time.time() - start
                hot_search_times.append(search_time)
                print(f"  Sorgu {i}: {search_time:.4f}s")
            
            avg_hot_search = sum(hot_search_times) / len(hot_search_times)
            self.results["milvus"]["avg_hot_search_time"] = avg_hot_search
            self.results["milvus"]["min_hot_search_time"] = min(hot_search_times)
            self.results["milvus"]["max_hot_search_time"] = max(hot_search_times)
            print(f"  Ortalama: {avg_hot_search:.4f}s")
            
            # Disk boyutu
            disk_size = os.path.getsize(db_path) / 1024 / 1024
            self.results["milvus"]["disk_size_mb"] = disk_size
            print(f"\n💾 Disk boyutu: {disk_size:.2f} MB")
            
            print("✅ Milvus benchmark tamamlandı")
            
        except Exception as e:
            print(f"❌ Hata: {e}")
            self.results["milvus"]["error"] = str(e)
    
    # ==================== Weaviate Benchmark ====================
    def benchmark_weaviate(self):
        """Weaviate performansını ölç"""
        print("\n" + "="*60)
        print("📊 Weaviate BENCHMARK")
        print("="*60)
        
        try:
            # Bağlantı zamanı
            start = time.time()
            client = weaviate.connect_to_local()
            connection_time = time.time() - start
            self.results["weaviate"]["connection_time"] = connection_time
            print(f"✓ Bağlantı zamanı: {connection_time:.4f}s")
            
            try:
                collection = client.collections.get("Documents")
                
                # Kayıt sayısı
                start = time.time()
                agg = collection.aggregate.over_all(total_count=True)
                count = agg.total_count
                count_time = time.time() - start
                self.results["weaviate"]["record_count"] = count
                self.results["weaviate"]["count_time"] = count_time
                print(f"✓ Toplam kayıt: {count} ({count_time:.4f}s)")
                
                # Soğuk arama (Vector)
                print("\n🔍 Soğuk başlangıç aramaları (Vector)...")
                cold_search_times = []
                for i, query in enumerate(self.test_queries, 1):
                    query_vector = self.model.encode(query).tolist()
                    start = time.time()
                    results = collection.query.near_vector(
                        near_vector=query_vector,
                        limit=5
                    )
                    search_time = time.time() - start
                    cold_search_times.append(search_time)
                    print(f"  Sorgu {i}: {search_time:.4f}s")
                
                avg_cold_search = sum(cold_search_times) / len(cold_search_times)
                self.results["weaviate"]["avg_cold_search_time"] = avg_cold_search
                self.results["weaviate"]["min_cold_search_time"] = min(cold_search_times)
                self.results["weaviate"]["max_cold_search_time"] = max(cold_search_times)
                print(f"  Ortalama: {avg_cold_search:.4f}s")
                
                # Sıcak arama (Vector)
                print("\n🔥 Sıcak aramalar (Vector)...")
                hot_search_times = []
                for i, query in enumerate(self.test_queries, 1):
                    query_vector = self.model.encode(query).tolist()
                    start = time.time()
                    results = collection.query.near_vector(
                        near_vector=query_vector,
                        limit=5
                    )
                    search_time = time.time() - start
                    hot_search_times.append(search_time)
                    print(f"  Sorgu {i}: {search_time:.4f}s")
                
                avg_hot_search = sum(hot_search_times) / len(hot_search_times)
                self.results["weaviate"]["avg_hot_search_time"] = avg_hot_search
                self.results["weaviate"]["min_hot_search_time"] = min(hot_search_times)
                self.results["weaviate"]["max_hot_search_time"] = max(hot_search_times)
                print(f"  Ortalama: {avg_hot_search:.4f}s")
                
                # Hybrid arama
                print("\n🔥 Hybrid arama (Vector + BM25)...")
                hybrid_search_times = []
                for i, query in enumerate(self.test_queries, 1):
                    query_vector = self.model.encode(query).tolist()
                    start = time.time()
                    results = collection.query.hybrid(
                        query=query,
                        vector=query_vector,
                        limit=5
                    )
                    search_time = time.time() - start
                    hybrid_search_times.append(search_time)
                    print(f"  Sorgu {i}: {search_time:.4f}s")
                
                avg_hybrid_search = sum(hybrid_search_times) / len(hybrid_search_times)
                self.results["weaviate"]["avg_hybrid_search_time"] = avg_hybrid_search
                print(f"  Ortalama: {avg_hybrid_search:.4f}s")
                
                print("✅ Weaviate benchmark tamamlandı")
                
            except Exception as e:
                print(f"⚠ Weaviate collection bulunamadı: {e}")
                self.results["weaviate"]["error"] = "Collection not found"
            
            client.close()
            
        except Exception as e:
            print(f"❌ Weaviate bağlantı hatası (sunucu çalışmıyor olabilir): {e}")
            self.results["weaviate"]["error"] = f"Connection failed: {str(e)}"
    
    # ==================== pgvector Benchmark ====================
    def benchmark_pgvector(self):
        """pgvector performansını ölç"""
        print("\n" + "="*60)
        print("📊 pgvector (PostgreSQL) BENCHMARK")
        print("="*60)
        
        try:
            # Bağlantı zamanı
            start = time.time()
            conn = psycopg2.connect(
                host="localhost",
                database="vector_db",
                user="postgres",
                password="yeni_sifre",
                port="5432"
            )
            connection_time = time.time() - start
            self.results["pgvector"]["connection_time"] = connection_time
            print(f"✓ Bağlantı zamanı: {connection_time:.4f}s")
            
            cursor = conn.cursor()
            
            # Kayıt sayısı
            start = time.time()
            cursor.execute("SELECT COUNT(*) FROM documents;")
            count = cursor.fetchone()[0]
            count_time = time.time() - start
            self.results["pgvector"]["record_count"] = count
            self.results["pgvector"]["count_time"] = count_time
            print(f"✓ Toplam kayıt: {count} ({count_time:.4f}s)")
            
            # Soğuk arama (Cosine)
            print("\n🔍 Soğuk başlangıç aramaları (Cosine)...")
            cold_search_times = []
            for i, query in enumerate(self.test_queries, 1):
                query_vector = self.model.encode(query).tolist()
                embedding_str = '[' + ','.join(map(str, query_vector)) + ']'
                
                start = time.time()
                cursor.execute(f"""
                    SELECT id, chunk_id, metin, 
                           embedding <=> '{embedding_str}'::vector AS distance
                    FROM documents
                    ORDER BY embedding <=> '{embedding_str}'::vector
                    LIMIT 5;
                """)
                results = cursor.fetchall()
                search_time = time.time() - start
                cold_search_times.append(search_time)
                print(f"  Sorgu {i}: {search_time:.4f}s")
            
            avg_cold_search = sum(cold_search_times) / len(cold_search_times)
            self.results["pgvector"]["avg_cold_search_time"] = avg_cold_search
            self.results["pgvector"]["min_cold_search_time"] = min(cold_search_times)
            self.results["pgvector"]["max_cold_search_time"] = max(cold_search_times)
            print(f"  Ortalama: {avg_cold_search:.4f}s")
            
            # Sıcak arama (Cosine)
            print("\n🔥 Sıcak aramalar (Cosine)...")
            hot_search_times = []
            for i, query in enumerate(self.test_queries, 1):
                query_vector = self.model.encode(query).tolist()
                embedding_str = '[' + ','.join(map(str, query_vector)) + ']'
                
                start = time.time()
                cursor.execute(f"""
                    SELECT id, chunk_id, metin, 
                           embedding <=> '{embedding_str}'::vector AS distance
                    FROM documents
                    ORDER BY embedding <=> '{embedding_str}'::vector
                    LIMIT 5;
                """)
                results = cursor.fetchall()
                search_time = time.time() - start
                hot_search_times.append(search_time)
                print(f"  Sorgu {i}: {search_time:.4f}s")
            
            avg_hot_search = sum(hot_search_times) / len(hot_search_times)
            self.results["pgvector"]["avg_hot_search_time"] = avg_hot_search
            self.results["pgvector"]["min_hot_search_time"] = min(hot_search_times)
            self.results["pgvector"]["max_hot_search_time"] = max(hot_search_times)
            print(f"  Ortalama: {avg_hot_search:.4f}s")
            
            cursor.close()
            conn.close()
            
            print("✅ pgvector benchmark tamamlandı")
            
        except Exception as e:
            print(f"❌ pgvector bağlantı hatası: {e}")
            self.results["pgvector"]["error"] = f"Connection failed: {str(e)}"
    
    def run_all_benchmarks(self):
        """Tüm benchmark'leri çalıştır"""
        print("\n" + "🚀"*30)
        print("VERITABANI PERFORMANS BENCHMARK'İ BAŞLIYOR")
        print("🚀"*30)
        
        self.benchmark_lancedb()
        self.benchmark_chromadb()
        self.benchmark_milvus()
        self.benchmark_weaviate()
        self.benchmark_pgvector()
        self.benchmark_qdrant()
        
        self.print_detailed_comparison()
        self.save_results()
    
    def print_detailed_comparison(self):
        """Detaylı karşılaştırmalı sonuçlar"""
        print("\n" + "="*60)
        print("📊 DETAYLI KARŞILAŞTIRMA")
        print("="*60)
        
        # Bağlantı Zamanı
        print("\n⚡ Bağlantı Zamanı:")
        connection_times = []
        for db_name, metrics in self.results.items():
            if "connection_time" in metrics and "error" not in metrics:
                time_val = metrics['connection_time']
                connection_times.append((db_name, time_val))
        
        connection_times.sort(key=lambda x: x[1])
        for i, (db_name, time_val) in enumerate(connection_times, 1):
            emoji = "🥇" if i == 1 else "🥈" if i == 2 else "🥉" if i == 3 else "  "
            print(f"{emoji} {i}. {db_name:12} : {time_val:.4f}s")
        
        # Soğuk Arama Zamanı
        print("\n🔍 Ortalama Soğuk Arama Zamanı:")
        cold_search_times = []
        for db_name, metrics in self.results.items():
            if "avg_cold_search_time" in metrics:
                time_val = metrics['avg_cold_search_time']
                cold_search_times.append((db_name, time_val))
        
        cold_search_times.sort(key=lambda x: x[1])
        for i, (db_name, time_val) in enumerate(cold_search_times, 1):
            emoji = "🥇" if i == 1 else "🥈" if i == 2 else "🥉" if i == 3 else "  "
            print(f"{emoji} {i}. {db_name:12} : {time_val:.4f}s")
        
        # Sıcak Arama Zamanı
        print("\n🔥 Ortalama Sıcak Arama Zamanı (Cached):")
        hot_search_times = []
        for db_name, metrics in self.results.items():
            if "avg_hot_search_time" in metrics:
                time_val = metrics['avg_hot_search_time']
                hot_search_times.append((db_name, time_val))
        
        hot_search_times.sort(key=lambda x: x[1])
        for i, (db_name, time_val) in enumerate(hot_search_times, 1):
            emoji = "🥇" if i == 1 else "🥈" if i == 2 else "🥉" if i == 3 else "  "
            print(f"{emoji} {i}. {db_name:12} : {time_val:.4f}s")
        
        # Disk Boyutu
        print("\n💾 Disk Boyutu:")
        disk_sizes = []
        for db_name, metrics in self.results.items():
            if "disk_size_mb" in metrics:
                size_val = metrics['disk_size_mb']
                disk_sizes.append((db_name, size_val))
        
        disk_sizes.sort(key=lambda x: x[1])
        for i, (db_name, size_val) in enumerate(disk_sizes, 1):
            emoji = "🥇" if i == 1 else "🥈" if i == 2 else "🥉" if i == 3 else "  "
            print(f"{emoji} {i}. {db_name:12} : {size_val:.2f} MB")
        
        # Kayıt Sayısı
        print("\n📈 Kayıt Sayısı:")
        for db_name, metrics in self.results.items():
            if "record_count" in metrics and "error" not in metrics:
                print(f"   {db_name:12} : {metrics['record_count']:,}")
        
        # Hata Durumu
        print("\n❌ Hata Durumu:")
        errors_found = False
        for db_name, metrics in self.results.items():
            if "error" in metrics:
                print(f"   {db_name:12} : {metrics['error']}")
                errors_found = True
        
        if not errors_found:
            print("   ✅ Tüm veritabanları başarılı")
        
        # En İyi Performans Özeti
        print("\n" + "="*60)
        print("🏆 EN İYİ PERFORMANS ÖZETİ")
        print("="*60)
        
        if connection_times:
            print(f"⚡ En Hızlı Bağlantı   : {connection_times[0][0]} ({connection_times[0][1]:.4f}s)")
        
        if cold_search_times:
            print(f"🔍 En Hızlı Soğuk Arama: {cold_search_times[0][0]} ({cold_search_times[0][1]:.4f}s)")
        
        if hot_search_times:
            print(f"🔥 En Hızlı Sıcak Arama: {hot_search_times[0][0]} ({hot_search_times[0][1]:.4f}s)")
        
        if disk_sizes:
            print(f"💾 En Küçük Disk       : {disk_sizes[0][0]} ({disk_sizes[0][1]:.2f} MB)")
    
    def save_results(self):
        """Sonuçları Excel dosyasına kaydet"""
        output_file = "/home/ugo/Documents/Python/bitirememe projesi/dataset_benchmark.xlsx"
        
        # Yeni workbook oluştur
        wb = openpyxl.Workbook()
        
        # Özet sayfası
        ws_summary = wb.active
        ws_summary.title = "Özet Karşılaştırma"
        
        # Stil tanımlamaları
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        subheader_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        title_font = Font(bold=True, size=14)
        center_align = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Başlık
        ws_summary.merge_cells('A1:H1')
        ws_summary['A1'] = "Vector Database Performance Benchmark"
        ws_summary['A1'].font = Font(bold=True, size=16)
        ws_summary['A1'].alignment = center_align
        
        ws_summary.merge_cells('A2:H2')
        ws_summary['A2'] = f"Tarih: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws_summary['A2'].alignment = center_align
        
        # Tüm veritabanlarını listele
        db_names = [db for db in self.results.keys() if "error" not in self.results[db]]
        
        # Ana tablo başlığı
        row = 4
        headers = ["Veritabanı", "Bağlantı (s)", "Kayıt Sayısı", "Soğuk Arama (s)", "Sıcak Arama (s)", "Min Arama (s)", "Max Arama (s)", "Disk (MB)"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws_summary.cell(row=row, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
        
        # Verileri doldur
        row += 1
        for db_name in db_names:
            metrics = self.results[db_name]
            
            ws_summary.cell(row=row, column=1, value=db_name)
            ws_summary.cell(row=row, column=2, value=round(metrics.get('connection_time', 0), 4))
            ws_summary.cell(row=row, column=3, value=metrics.get('record_count', 0))
            ws_summary.cell(row=row, column=4, value=round(metrics.get('avg_cold_search_time', 0), 4))
            ws_summary.cell(row=row, column=5, value=round(metrics.get('avg_hot_search_time', 0), 4))
            ws_summary.cell(row=row, column=6, value=round(metrics.get('min_hot_search_time', 0), 4))
            ws_summary.cell(row=row, column=7, value=round(metrics.get('max_hot_search_time', 0), 4))
            ws_summary.cell(row=row, column=8, value=round(metrics.get('disk_size_mb', 0), 2))
            
            # Stil uygula
            for col_idx in range(1, 9):
                cell = ws_summary.cell(row=row, column=col_idx)
                cell.border = border
                cell.alignment = center_align
            
            row += 1
        
        # Sıralama tabloları
        row += 2
        
        # En İyi Performanslar
        ws_summary.merge_cells(f'A{row}:D{row}')
        ws_summary[f'A{row}'] = "🏆 EN İYİ PERFORMANSLAR"
        ws_summary[f'A{row}'].font = title_font
        ws_summary[f'A{row}'].alignment = center_align
        row += 1
        
        # Bağlantı Hızı
        ws_summary[f'A{row}'] = "Sıra"
        ws_summary[f'B{row}'] = "⚡ En Hızlı Bağlantı"
        ws_summary[f'C{row}'] = "Süre (s)"
        ws_summary[f'D{row}'] = "Fark"
        for col in ['A', 'B', 'C', 'D']:
            ws_summary[f'{col}{row}'].fill = subheader_fill
            ws_summary[f'{col}{row}'].font = Font(bold=True)
            ws_summary[f'{col}{row}'].border = border
            ws_summary[f'{col}{row}'].alignment = center_align
        
        row += 1
        connection_times = [(db, m['connection_time']) for db, m in self.results.items() 
                           if 'connection_time' in m and 'error' not in m]
        connection_times.sort(key=lambda x: x[1])
        
        for i, (db_name, time_val) in enumerate(connection_times[:5], 1):
            emoji = "🥇" if i == 1 else "🥈" if i == 2 else "🥉" if i == 3 else f"{i}."
            ws_summary[f'A{row}'] = emoji
            ws_summary[f'B{row}'] = db_name
            ws_summary[f'C{row}'] = round(time_val, 4)
            if i == 1:
                ws_summary[f'D{row}'] = "En Hızlı"
            else:
                diff_percent = ((time_val / connection_times[0][1]) - 1) * 100
                ws_summary[f'D{row}'] = f"+%{diff_percent:.1f}"
            
            for col in ['A', 'B', 'C', 'D']:
                ws_summary[f'{col}{row}'].border = border
                ws_summary[f'{col}{row}'].alignment = center_align
            row += 1
        
        # Soğuk Arama
        row += 1
        ws_summary[f'A{row}'] = "Sıra"
        ws_summary[f'B{row}'] = "🔍 En Hızlı Soğuk Arama"
        ws_summary[f'C{row}'] = "Süre (s)"
        ws_summary[f'D{row}'] = "Fark"
        for col in ['A', 'B', 'C', 'D']:
            ws_summary[f'{col}{row}'].fill = subheader_fill
            ws_summary[f'{col}{row}'].font = Font(bold=True)
            ws_summary[f'{col}{row}'].border = border
            ws_summary[f'{col}{row}'].alignment = center_align
        
        row += 1
        cold_times = [(db, m['avg_cold_search_time']) for db, m in self.results.items() 
                     if 'avg_cold_search_time' in m]
        cold_times.sort(key=lambda x: x[1])
        
        for i, (db_name, time_val) in enumerate(cold_times[:5], 1):
            emoji = "🥇" if i == 1 else "🥈" if i == 2 else "🥉" if i == 3 else f"{i}."
            ws_summary[f'A{row}'] = emoji
            ws_summary[f'B{row}'] = db_name
            ws_summary[f'C{row}'] = round(time_val, 4)
            if i == 1:
                ws_summary[f'D{row}'] = "En Hızlı"
            else:
                diff_percent = ((time_val / cold_times[0][1]) - 1) * 100
                ws_summary[f'D{row}'] = f"+%{diff_percent:.1f}"
            
            for col in ['A', 'B', 'C', 'D']:
                ws_summary[f'{col}{row}'].border = border
                ws_summary[f'{col}{row}'].alignment = center_align
            row += 1
        
        # Sıcak Arama
        row += 1
        ws_summary[f'A{row}'] = "Sıra"
        ws_summary[f'B{row}'] = "🔥 En Hızlı Sıcak Arama"
        ws_summary[f'C{row}'] = "Süre (s)"
        ws_summary[f'D{row}'] = "Fark"
        for col in ['A', 'B', 'C', 'D']:
            ws_summary[f'{col}{row}'].fill = subheader_fill
            ws_summary[f'{col}{row}'].font = Font(bold=True)
            ws_summary[f'{col}{row}'].border = border
            ws_summary[f'{col}{row}'].alignment = center_align
        
        row += 1
        hot_times = [(db, m['avg_hot_search_time']) for db, m in self.results.items() 
                    if 'avg_hot_search_time' in m]
        hot_times.sort(key=lambda x: x[1])
        
        for i, (db_name, time_val) in enumerate(hot_times[:5], 1):
            emoji = "🥇" if i == 1 else "🥈" if i == 2 else "🥉" if i == 3 else f"{i}."
            ws_summary[f'A{row}'] = emoji
            ws_summary[f'B{row}'] = db_name
            ws_summary[f'C{row}'] = round(time_val, 4)
            if i == 1:
                ws_summary[f'D{row}'] = "En Hızlı"
            else:
                diff_percent = ((time_val / hot_times[0][1]) - 1) * 100
                ws_summary[f'D{row}'] = f"+%{diff_percent:.1f}"
            
            for col in ['A', 'B', 'C', 'D']:
                ws_summary[f'{col}{row}'].border = border
                ws_summary[f'{col}{row}'].alignment = center_align
            row += 1
        
        # Disk Boyutu
        row += 1
        ws_summary[f'A{row}'] = "Sıra"
        ws_summary[f'B{row}'] = "💾 En Küçük Disk Kullanımı"
        ws_summary[f'C{row}'] = "Boyut (MB)"
        ws_summary[f'D{row}'] = "Fark"
        for col in ['A', 'B', 'C', 'D']:
            ws_summary[f'{col}{row}'].fill = subheader_fill
            ws_summary[f'{col}{row}'].font = Font(bold=True)
            ws_summary[f'{col}{row}'].border = border
            ws_summary[f'{col}{row}'].alignment = center_align
        
        row += 1
        disk_sizes = [(db, m['disk_size_mb']) for db, m in self.results.items() 
                     if 'disk_size_mb' in m]
        disk_sizes.sort(key=lambda x: x[1])
        
        for i, (db_name, size_val) in enumerate(disk_sizes[:5], 1):
            emoji = "🥇" if i == 1 else "🥈" if i == 2 else "🥉" if i == 3 else f"{i}."
            ws_summary[f'A{row}'] = emoji
            ws_summary[f'B{row}'] = db_name
            ws_summary[f'C{row}'] = round(size_val, 2)
            if i == 1:
                ws_summary[f'D{row}'] = "En Küçük"
            else:
                diff_percent = ((size_val / disk_sizes[0][1]) - 1) * 100
                ws_summary[f'D{row}'] = f"+%{diff_percent:.1f}"
            
            for col in ['A', 'B', 'C', 'D']:
                ws_summary[f'{col}{row}'].border = border
                ws_summary[f'{col}{row}'].alignment = center_align
            row += 1
        
        # Sütun genişliklerini ayarla
        ws_summary.column_dimensions['A'].width = 12
        ws_summary.column_dimensions['B'].width = 25
        ws_summary.column_dimensions['C'].width = 15
        ws_summary.column_dimensions['D'].width = 15
        ws_summary.column_dimensions['E'].width = 15
        ws_summary.column_dimensions['F'].width = 15
        ws_summary.column_dimensions['G'].width = 15
        ws_summary.column_dimensions['H'].width = 15
        
        # Detaylı Sonuçlar Sayfası
        ws_details = wb.create_sheet("Tüm Metrikler")
        
        row = 1
        ws_details['A1'] = "Veritabanı"
        ws_details['B1'] = "Metrik"
        ws_details['C1'] = "Değer"
        ws_details['D1'] = "Birim"
        for col in ['A', 'B', 'C', 'D']:
            ws_details[f'{col}1'].fill = header_fill
            ws_details[f'{col}1'].font = header_font
            ws_details[f'{col}1'].border = border
            ws_details[f'{col}1'].alignment = center_align
        
        row = 2
        for db_name, metrics in self.results.items():
            for metric_name, metric_value in metrics.items():
                ws_details[f'A{row}'] = db_name
                ws_details[f'B{row}'] = metric_name
                
                if isinstance(metric_value, float):
                    ws_details[f'C{row}'] = round(metric_value, 4)
                else:
                    ws_details[f'C{row}'] = metric_value
                
                # Birim ekle
                if 'time' in metric_name.lower():
                    ws_details[f'D{row}'] = "saniye"
                elif 'size' in metric_name.lower() or 'mb' in metric_name.lower():
                    ws_details[f'D{row}'] = "MB"
                elif 'count' in metric_name.lower():
                    ws_details[f'D{row}'] = "adet"
                else:
                    ws_details[f'D{row}'] = "-"
                
                for col in ['A', 'B', 'C', 'D']:
                    ws_details[f'{col}{row}'].border = border
                    ws_details[f'{col}{row}'].alignment = center_align
                row += 1
        
        ws_details.column_dimensions['A'].width = 20
        ws_details.column_dimensions['B'].width = 30
        ws_details.column_dimensions['C'].width = 20
        ws_details.column_dimensions['D'].width = 15
        
        # Kaydet
        wb.save(output_file)
        print(f"\n💾 Detaylı sonuçlar Excel dosyasına kaydedildi: {output_file}")
        print(f"   📊 Özet Karşılaştırma: Tablo formatında karşılaştırmalı sonuçlar")
        print(f"   📋 Tüm Metrikler: Detaylı tüm ölçümler")
        
        # JSON'a da kaydet (yedek)
        json_file = output_file.replace('.xlsx', '.json')
        with open(json_file, 'w', encoding='utf-8') as f:
            json.dump(self.results, f, indent=2, ensure_ascii=False)
        print(f"   💾 JSON yedeği: {json_file}")

if __name__ == "__main__":
    benchmark = DatabaseBenchmark()
    benchmark.run_all_benchmarks()